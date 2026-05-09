import io
import json
from collections import Counter, defaultdict
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q, Count, Prefetch, F
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.views.decorators.http import require_POST
from .models import (
    Osoba, Grob, Sektor, Zdjecie, Relacja, Zgloszenie, Profil, HistoriaZmian,
    Wspomnienie, Swieca, ZapisaneSzukanie, Wpis,
    Tag, Panorama, HotspotPanoramy, SubskrypcjaPush, TokenLogowania, Komentarz,
    Trasa, TrasaPunkt, Odznaka, UzytkownikOdznaka, Newsletter,
    Kwiat, Nagranie, GlosNagrobek, IntencjaMszalna, Zaproszenie, GeoCache,
    List, PytanieQuiz, WatekForum, PostForum, Webhook,
    WyszukiwanieLog, ZdjecieDronowe, KonkursFoto, ZgloszenieKonkursowe, GlosKonkursowy,
    TagWpisu, PlanZwiedzania, OdwiedzinyOsoba, FeaturedTygodnia,
    Powiadomienie, OpiekunGrobu, PrywatnaNotatka, HasloSlownik,
    EtykietaOsoby, WydarzenieParafialne,
    Sonda, OdpowiedzSondy, GlosSondy, Kondolencja, ZbiorkaRenowacja, NotkaCmentarna,
)


SZYDLOW_CENTRUM = (50.5847, 20.8327)


def _antybot(request):
    """Honeypot + minimalny czas wypełnienia. Zwraca True jeśli to bot."""
    import time
    if (request.POST.get('_pulapka') or '').strip():
        return True
    try:
        ts = float(request.POST.get('_ts', '0'))
        if time.time() - ts < 2.0:
            return True
    except (ValueError, TypeError):
        return True
    return False


def _pole_antybot_html():
    import time
    return (
        f'<input type="text" name="_pulapka" tabindex="-1" autocomplete="off" '
        f'style="position:absolute;left:-9999px;width:1px;height:1px;opacity:0;" aria-hidden="true">'
        f'<input type="hidden" name="_ts" value="{time.time():.0f}">'
    )


def _szukaj_fts(osoby_qs, query):
    """SQLite FTS5 z fallbackiem na icontains. Wspiera prefiksy (Kowal* -> Kowalski)."""
    from django.db import connection
    try:
        with connection.cursor() as cur:
            tokeny = [t for t in query.replace('"', '').split() if t]
            if not tokeny:
                return osoby_qs.none()
            fts_q = ' '.join(t + '*' for t in tokeny)
            cur.execute('SELECT rowid FROM osoba_fts WHERE osoba_fts MATCH %s LIMIT 5000', [fts_q])
            ids = [r[0] for r in cur.fetchall()]
        if ids:
            return osoby_qs.filter(pk__in=ids)
        return osoby_qs.none()
    except Exception:
        return osoby_qs.filter(
            Q(nazwisko__icontains=query)
            | Q(imie__icontains=query)
            | Q(nazwisko_rodowe__icontains=query)
        )


from django.views.decorators.cache import cache_page


@cache_page(60 * 15)
def home(request):
    context = {
        'liczba_osob': Osoba.objects.count(),
        'liczba_grobow': Grob.objects.count(),
        'liczba_sektorow': Sektor.objects.count(),
    }
    return render(request, 'groby/home.html', context)


def szukaj(request):
    query = request.GET.get('q', '').strip()
    sektor_id = request.GET.get('sektor', '').strip()
    typ = request.GET.get('typ', '').strip()
    rok_od = request.GET.get('rok_od', '').strip()
    rok_do = request.GET.get('rok_do', '').strip()

    osoby = Osoba.objects.select_related('grob', 'grob__sektor')

    if query:
        osoby = _szukaj_fts(osoby, query)
    if sektor_id:
        osoby = osoby.filter(grob__sektor_id=sektor_id)
    if typ:
        osoby = osoby.filter(grob__typ=typ)
    if rok_od.isdigit():
        osoby = osoby.filter(data_smierci__year__gte=int(rok_od))
    if rok_do.isdigit():
        osoby = osoby.filter(data_smierci__year__lte=int(rok_do))

    tag_slug = request.GET.get('tag', '').strip()
    if tag_slug:
        osoby = osoby.filter(tagi__slug=tag_slug)

    osoby = osoby.order_by('nazwisko', 'imie').distinct()
    total = osoby.count()

    sugestia = None
    if query and total == 0:
        sugestia = _zaproponuj_nazwisko(query)

    if query:
        try:
            WyszukiwanieLog.objects.create(
                fraza=query[:200],
                user=request.user if request.user.is_authenticated else None,
                ip_hash=_hash_ip(request),
                liczba_wynikow=total,
            )
        except Exception:
            pass

    paginator = Paginator(osoby, 20)
    page = paginator.get_page(request.GET.get('page'))

    aktywne_filtry = bool(query or sektor_id or typ or rok_od or rok_do)

    params = request.GET.copy()
    params.pop('page', None)
    querystring_bez_page = params.urlencode()

    context = {
        'query': query,
        'sektor_id': sektor_id,
        'typ': typ,
        'rok_od': rok_od,
        'rok_do': rok_do,
        'page': page,
        'total': total,
        'sektory': Sektor.objects.order_by('nazwa'),
        'typy': Grob.TYP_CHOICES,
        'aktywne_filtry': aktywne_filtry,
        'querystring_bez_page': querystring_bez_page,
        'sugestia': sugestia,
        'tagi': Tag.objects.all(),
        'aktywny_tag': tag_slug,
    }
    return render(request, 'groby/szukaj.html', context)


def _zaproponuj_nazwisko(query):
    """Najbliższe nazwisko/imie do zapytania, używając difflib."""
    import difflib
    fraza = query.strip().lower()
    if len(fraza) < 3:
        return None
    nazwiska = set()
    for n, i in Osoba.objects.values_list('nazwisko', 'imie'):
        if n: nazwiska.add(n)
        if i: nazwiska.add(i)
    bliskie = difflib.get_close_matches(fraza, [n.lower() for n in nazwiska], n=1, cutoff=0.7)
    if not bliskie:
        return None
    bliska_lower = bliskie[0]
    for n in nazwiska:
        if n.lower() == bliska_lower:
            return n
    return None


@cache_page(60 * 15)
def sektory_list(request):
    sektory = Sektor.objects.annotate(
        liczba_grobow=Count('groby', distinct=True),
        liczba_osob=Count('groby__osoby', distinct=True),
    ).order_by('nazwa')
    return render(request, 'groby/sektory_list.html', {'sektory': sektory})


def sektor_detail(request, pk):
    sektor = get_object_or_404(Sektor, pk=pk)
    groby = sektor.groby.prefetch_related('osoby').order_by('numer')
    paginator = Paginator(groby, 30)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'groby/sektor_detail.html', {'sektor': sektor, 'page': page})


def grob_detail(request, pk):
    grob = get_object_or_404(
        Grob.objects.select_related('sektor').prefetch_related('osoby', 'zdjecia'),
        pk=pk,
    )
    obserwuje = False
    if request.user.is_authenticated:
        obserwuje = grob.obserwujacy.filter(user=request.user).exists()
    return render(request, 'groby/grob_detail.html', {
        'grob': grob,
        'zdjecia': grob.zdjecia.all(),
        'obserwuje': obserwuje,
    })


def o_cmentarzu(request):
    return render(request, 'groby/o_cmentarzu.html')


def sugestie(request):
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})
    osoby = (
        Osoba.objects
        .filter(Q(nazwisko__istartswith=q) | Q(imie__istartswith=q) | Q(nazwisko_rodowe__istartswith=q))
        .select_related('grob', 'grob__sektor')
        .order_by('nazwisko', 'imie')[:8]
    )
    data = []
    for o in osoby:
        rodowe = f' (z d. {o.nazwisko_rodowe})' if o.nazwisko_rodowe else ''
        rok_ur = o.data_urodzenia.year if o.data_urodzenia else '?'
        rok_sm = o.data_smierci.year if o.data_smierci else '?'
        data.append({
            'nazwa': f'{o.imie} {o.nazwisko}{rodowe}',
            'lata': f'{rok_ur} — {rok_sm}',
            'grob': f'Sektor {o.grob.sektor.nazwa} · nr {o.grob.numer}',
            'url': reverse('groby:osoba_detail', args=[o.pk]),
        })
    return JsonResponse({'results': data})


def statystyki(request):
    typy_labels = dict(Grob.TYP_CHOICES)
    typy_qs = Grob.objects.values('typ').annotate(c=Count('id')).order_by('-c')
    typy = [{'label': typy_labels.get(t['typ'], t['typ']), 'value': t['c']} for t in typy_qs]

    sektor_qs = Sektor.objects.annotate(liczba=Count('groby')).values('nazwa', 'liczba').order_by('nazwa')
    sektory_dane = [{'label': f"Sektor {s['nazwa']}", 'value': s['liczba']} for s in sektor_qs]

    dekady_counter = Counter()
    for ds in Osoba.objects.filter(data_smierci__isnull=False).values_list('data_smierci', flat=True):
        dekady_counter[(ds.year // 10) * 10] += 1
    dekady = [{'label': f'{d}.', 'value': v} for d, v in sorted(dekady_counter.items())]

    wieki = []
    for o in Osoba.objects.filter(data_urodzenia__isnull=False, data_smierci__isnull=False):
        w = o.wiek
        if w is not None and w >= 0:
            wieki.append(w)

    sredni_wiek = round(sum(wieki) / len(wieki), 1) if wieki else 0
    najstarsza = max(wieki) if wieki else 0
    najmlodsza = min(wieki) if wieki else 0

    osob_na_grob_counter = Counter()
    for g in Grob.objects.annotate(lo=Count('osoby')).values_list('lo', flat=True):
        osob_na_grob_counter[g] += 1
    osob_na_grob = [{'label': f'{k} osoba' if k == 1 else f'{k} osób', 'value': v} for k, v in sorted(osob_na_grob_counter.items())]

    heatmapa = defaultdict(lambda: [0]*12)
    for ds in Osoba.objects.filter(data_smierci__isnull=False).values_list('data_smierci', flat=True):
        dekada = (ds.year // 10) * 10
        heatmapa[dekada][ds.month - 1] += 1
    heatmapa_dane = [{'dekada': f'{d}.', 'wartosci': heatmapa[d]} for d in sorted(heatmapa)]
    miesiace = ['Sty', 'Lut', 'Mar', 'Kwi', 'Maj', 'Cze', 'Lip', 'Sie', 'Wrz', 'Paź', 'Lis', 'Gru']

    context = {
        'typy_json': json.dumps(typy, ensure_ascii=False),
        'sektory_json': json.dumps(sektory_dane, ensure_ascii=False),
        'dekady_json': json.dumps(dekady, ensure_ascii=False),
        'osob_na_grob_json': json.dumps(osob_na_grob, ensure_ascii=False),
        'heatmapa_json': json.dumps(heatmapa_dane, ensure_ascii=False),
        'miesiace_json': json.dumps(miesiace, ensure_ascii=False),
        'sredni_wiek': sredni_wiek,
        'najstarsza': najstarsza,
        'najmlodsza': najmlodsza,
        'liczba_osob': Osoba.objects.count(),
        'liczba_grobow': Grob.objects.count(),
        'liczba_sektorow': Sektor.objects.count(),
    }
    return render(request, 'groby/statystyki.html', context)


def mapa(request):
    from django.conf import settings
    from PIL import Image

    Image.MAX_IMAGE_PIXELS = None

    sektor_id = request.GET.get('sektor', '').strip()
    typ = request.GET.get('typ', '').strip()
    rok_od = request.GET.get('rok_od', '').strip()
    rok_do = request.GET.get('rok_do', '').strip()

    groby = Grob.objects.filter(
        plan_x__isnull=False,
        plan_y__isnull=False,
    ).select_related('sektor').prefetch_related('osoby').distinct()

    if sektor_id:
        groby = groby.filter(sektor_id=sektor_id)
    if typ:
        groby = groby.filter(typ=typ)
    if rok_od.isdigit():
        groby = groby.filter(osoby__data_smierci__year__gte=int(rok_od))
    if rok_do.isdigit():
        groby = groby.filter(osoby__data_smierci__year__lte=int(rok_do))

    dane = []
    for g in groby:
        osoby_str = [
            f'{o.imie} {o.nazwisko}' + (f' (z d. {o.nazwisko_rodowe})' if o.nazwisko_rodowe else '')
            for o in g.osoby.all()
        ]
        dane.append({
            'id': g.pk,
            'x': g.plan_x,
            'y': g.plan_y,
            'sektor': g.sektor.nazwa,
            'rzad': g.rzad,
            'numer': g.numer,
            'typ': g.get_typ_display(),
            'osoby': osoby_str,
            'url': reverse('groby:grob_detail', args=[g.pk]),
        })

    plan_url = settings.MEDIA_URL + settings.PLAN_IMAGE if settings.PLAN_IMAGE else None
    plan_w = plan_h = 0
    if plan_url:
        try:
            sciezka = settings.MEDIA_ROOT / settings.PLAN_IMAGE
            with Image.open(sciezka) as im:
                plan_w, plan_h = im.size
        except (FileNotFoundError, OSError):
            plan_url = None

    tryb_edycji = request.user.is_authenticated and request.user.is_staff

    context = {
        'groby_json': json.dumps(dane, ensure_ascii=False),
        'liczba': len(dane),
        'liczba_wszystkich': Grob.objects.count(),
        'plan_url': plan_url,
        'plan_w': plan_w,
        'plan_h': plan_h,
        'tryb_edycji': tryb_edycji,
        'sektory': Sektor.objects.order_by('nazwa'),
        'typy': Grob.TYP_CHOICES,
        'sektor_id': sektor_id,
        'typ': typ,
        'rok_od': rok_od,
        'rok_do': rok_do,
        'aktywne_filtry': bool(sektor_id or typ or rok_od or rok_do),
    }
    return render(request, 'groby/mapa.html', context)


def zapisz_pozycje(request):
    from django.http import JsonResponse, HttpResponseNotAllowed
    from django.contrib.auth.decorators import login_required

    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    if not (request.user.is_authenticated and request.user.is_staff):
        return JsonResponse({'ok': False, 'error': 'Brak uprawnień'}, status=403)
    try:
        payload = json.loads(request.body or '{}')
        grob_id = int(payload.get('grob_id'))
        x = float(payload.get('x'))
        y = float(payload.get('y'))
    except (ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({'ok': False, 'error': 'Nieprawidłowe dane'}, status=400)

    try:
        grob = Grob.objects.select_related('sektor').get(pk=grob_id)
    except Grob.DoesNotExist:
        return JsonResponse({'ok': False, 'error': f'Grób {grob_id} nie istnieje'}, status=404)

    grob.plan_x = x
    grob.plan_y = y
    grob.save(update_fields=['plan_x', 'plan_y', 'data_modyfikacji'])
    tytul = f'{grob.sektor.nazwa}/{grob.rzad}/{grob.numer}' if grob.rzad else f'{grob.sektor.nazwa}/{grob.numer}'
    return JsonResponse({'ok': True, 'tytul': tytul, 'id': grob.id})


def osoba_detail(request, pk):
    from datetime import timedelta
    from django.utils import timezone
    osoba = get_object_or_404(Osoba.objects.select_related('grob', 'grob__sektor'), pk=pk)
    relacje = []
    for r in Relacja.objects.filter(Q(osoba_a=osoba) | Q(osoba_b=osoba)).select_related('osoba_a', 'osoba_b', 'osoba_a__grob__sektor', 'osoba_b__grob__sektor'):
        if r.osoba_a_id == osoba.pk:
            druga, kierunek = r.osoba_b, r.typ
        else:
            druga, kierunek = r.osoba_a, r.typ + '_zwrotna'
        relacje.append({'osoba': druga, 'typ': r.typ, 'kierunek': kierunek, 'etykieta': r.get_typ_display()})
    obserwuje = False
    if request.user.is_authenticated:
        obserwuje = osoba.obserwujacy.filter(user=request.user).exists()
    granica = timezone.now() - timedelta(hours=24)
    aktywne_swiece = osoba.swiece.filter(data_zapalenia__gte=granica)
    wspomnienia = osoba.wspomnienia.filter(status='zaakceptowane').prefetch_related(
        'komentarze__autor_user',
    )
    today = timezone.localdate()
    OdwiedzinyOsoba.objects.get_or_create(osoba=osoba, data=today)
    OdwiedzinyOsoba.objects.filter(osoba=osoba, data=today).update(licznik=F('licznik') + 1)
    licznik_total = sum(OdwiedzinyOsoba.objects.filter(osoba=osoba).values_list('licznik', flat=True))
    sparkline = list(OdwiedzinyOsoba.objects.filter(osoba=osoba).order_by('-data')[:30])
    sparkline.reverse()
    return render(request, 'groby/osoba_detail.html', {
        'osoba': osoba,
        'relacje': relacje,
        'obserwuje': obserwuje,
        'wspomnienia': wspomnienia,
        'liczba_swiec': aktywne_swiece.count(),
        'ostatnie_swiece': aktywne_swiece.order_by('-data_zapalenia')[:8],
        'licznik_odwiedzin': licznik_total,
        'sparkline': sparkline,
    })


def _hash_ip(request):
    import hashlib
    ip = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get('REMOTE_ADDR', '')
    return hashlib.sha256(ip.encode()).hexdigest()[:64]


@require_POST
def zapal_swiece(request, pk):
    from datetime import timedelta
    from django.utils import timezone
    osoba = get_object_or_404(Osoba, pk=pk)
    if _antybot(request):
        return redirect('groby:osoba_detail', pk=pk)
    granica = timezone.now() - timedelta(minutes=10)
    iph = _hash_ip(request)
    if Swieca.objects.filter(osoba=osoba, ip_hash=iph, data_zapalenia__gte=granica).exists():
        messages.info(request, 'Już zapaliłeś świeczkę dla tej osoby. Spróbuj za chwilę.')
        return redirect('groby:osoba_detail', pk=pk)
    Swieca.objects.create(
        osoba=osoba,
        autor_user=request.user if request.user.is_authenticated else None,
        intencja=(request.POST.get('intencja') or '').strip()[:200],
        ip_hash=iph,
    )
    messages.success(request, 'Świeczka zapalona — pali się przez 24 godziny.')
    return redirect('groby:osoba_detail', pk=pk)


@require_POST
def dodaj_wspomnienie(request, pk):
    osoba = get_object_or_404(Osoba, pk=pk)
    if _antybot(request):
        messages.error(request, 'Wykryto podejrzaną aktywność.')
        return redirect('groby:osoba_detail', pk=pk)
    tresc = (request.POST.get('tresc') or '').strip()
    if not tresc:
        messages.error(request, 'Treść wspomnienia nie może być pusta.')
        return redirect('groby:osoba_detail', pk=pk)
    Wspomnienie.objects.create(
        osoba=osoba,
        autor_user=request.user if request.user.is_authenticated else None,
        autor_imie=(request.POST.get('autor_imie') or '').strip()[:100],
        tresc=tresc[:5000],
    )
    messages.success(request, 'Dziękujemy. Wspomnienie zostanie wyświetlone po akceptacji moderatora.')
    return redirect('groby:osoba_detail', pk=pk)


def kalendarz_rocznic(request):
    from datetime import date, timedelta
    dzis = date.today()
    okno = int(request.GET.get('dni', '30'))
    okno = max(7, min(okno, 365))
    osoby = Osoba.objects.exclude(data_smierci__isnull=True).select_related('grob__sektor')
    rocznice = []
    for o in osoby:
        try:
            rocznica = o.data_smierci.replace(year=dzis.year)
        except ValueError:
            rocznica = o.data_smierci.replace(year=dzis.year, day=28)
        if rocznica < dzis:
            try:
                rocznica = o.data_smierci.replace(year=dzis.year + 1)
            except ValueError:
                rocznica = o.data_smierci.replace(year=dzis.year + 1, day=28)
        delta = (rocznica - dzis).days
        if delta <= okno:
            lat = rocznica.year - o.data_smierci.year
            rocznice.append({
                'osoba': o,
                'data': rocznica,
                'za_dni': delta,
                'lat_temu': lat,
                'oryginal': o.data_smierci,
            })
    rocznice.sort(key=lambda r: r['za_dni'])
    return render(request, 'groby/kalendarz.html', {
        'rocznice': rocznice,
        'okno': okno,
        'dzis': dzis,
    })


def kalendarz_ical(request):
    from datetime import date, datetime, timedelta
    from django.utils import timezone
    dzis = date.today()
    okno = int(request.GET.get('dni', '90'))
    okno = max(7, min(okno, 365))
    osoby = Osoba.objects.exclude(data_smierci__isnull=True).select_related('grob__sektor')
    linie = [
        'BEGIN:VCALENDAR',
        'VERSION:2.0',
        'PRODID:-//Informator Cmentarny Szydlow//PL',
        'CALSCALE:GREGORIAN',
        'METHOD:PUBLISH',
        'X-WR-CALNAME:Rocznice śmierci — cmentarz w Szydłowie',
    ]
    for o in osoby:
        try:
            rocznica = o.data_smierci.replace(year=dzis.year)
        except ValueError:
            rocznica = o.data_smierci.replace(year=dzis.year, day=28)
        if rocznica < dzis:
            try:
                rocznica = o.data_smierci.replace(year=dzis.year + 1)
            except ValueError:
                rocznica = o.data_smierci.replace(year=dzis.year + 1, day=28)
        if (rocznica - dzis).days > okno:
            continue
        lat = rocznica.year - o.data_smierci.year
        uid = f'osoba-{o.pk}-{rocznica.strftime("%Y%m%d")}@cmentarz.szydlow'
        opis = f'{lat}. rocznica śmierci. Sektor {o.grob.sektor.nazwa}, grób {o.grob.numer}.'
        linie += [
            'BEGIN:VEVENT',
            f'UID:{uid}',
            f'DTSTAMP:{timezone.now().strftime("%Y%m%dT%H%M%SZ")}',
            f'DTSTART;VALUE=DATE:{rocznica.strftime("%Y%m%d")}',
            f'DTEND;VALUE=DATE:{(rocznica + timedelta(days=1)).strftime("%Y%m%d")}',
            f'SUMMARY:Rocznica śmierci — {o.imie} {o.nazwisko}',
            f'DESCRIPTION:{opis}',
            'END:VEVENT',
        ]
    linie.append('END:VCALENDAR')
    resp = HttpResponse('\r\n'.join(linie), content_type='text/calendar; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="rocznice.ics"'
    return resp


def drzewo(request, pk):
    osoba = get_object_or_404(Osoba.objects.select_related('grob__sektor'), pk=pk)
    rodzice, dzieci, malzonkowie, rodzenstwo = [], [], [], []
    for r in Relacja.objects.filter(Q(osoba_a=osoba) | Q(osoba_b=osoba)).select_related('osoba_a', 'osoba_b'):
        if r.typ == 'rodzic':
            if r.osoba_a_id == osoba.pk:
                dzieci.append(r.osoba_b)
            else:
                rodzice.append(r.osoba_a)
        elif r.typ == 'malzenstwo':
            malzonkowie.append(r.osoba_b if r.osoba_a_id == osoba.pk else r.osoba_a)
        elif r.typ == 'rodzenstwo':
            rodzenstwo.append(r.osoba_b if r.osoba_a_id == osoba.pk else r.osoba_a)
    return render(request, 'groby/drzewo.html', {
        'osoba': osoba,
        'rodzice': rodzice,
        'dzieci': dzieci,
        'malzonkowie': malzonkowie,
        'rodzenstwo': rodzenstwo,
    })


def eksport_csv(request):
    import csv
    osoby = _filtruj_osoby(request)[:5000]
    resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    resp['Content-Disposition'] = 'attachment; filename="osoby.csv"'
    resp.write('﻿')
    w = csv.writer(resp, delimiter=';')
    w.writerow(['Nazwisko', 'Imię', 'Drugie imię', 'Nazwisko rodowe', 'Data urodzenia', 'Data śmierci', 'Wiek', 'Sektor', 'Rząd', 'Numer grobu', 'Typ grobu'])
    for o in osoby:
        w.writerow([
            o.nazwisko, o.imie, o.drugie_imie, o.nazwisko_rodowe,
            o.data_urodzenia or '', o.data_smierci or '', o.wiek or '',
            o.grob.sektor.nazwa, o.grob.rzad, o.grob.numer, o.grob.get_typ_display(),
        ])
    return resp


def eksport_xlsx(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    osoby = _filtruj_osoby(request)[:10000]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Osoby'
    naglowek = ['Nazwisko', 'Imię', 'Drugie imię', 'Nazwisko rodowe', 'Data urodzenia', 'Data śmierci', 'Wiek', 'Sektor', 'Rząd', 'Numer grobu', 'Typ grobu']
    ws.append(naglowek)
    fill = PatternFill(start_color='2E4430', end_color='2E4430', fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for c in ws[1]:
        c.fill, c.font, c.alignment = fill, font, Alignment(horizontal='center')
    for o in osoby:
        ws.append([
            o.nazwisko, o.imie, o.drugie_imie, o.nazwisko_rodowe,
            o.data_urodzenia, o.data_smierci, o.wiek,
            o.grob.sektor.nazwa, o.grob.rzad, o.grob.numer, o.grob.get_typ_display(),
        ])
    for col_idx, szer in enumerate([18, 14, 12, 16, 13, 13, 6, 8, 6, 10, 14], start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = szer
    ws.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="osoby.xlsx"'
    return resp


def _filtruj_osoby(request):
    osoby = Osoba.objects.select_related('grob', 'grob__sektor')
    query = request.GET.get('q', '').strip()
    sektor_id = request.GET.get('sektor', '').strip()
    typ = request.GET.get('typ', '').strip()
    rok_od = request.GET.get('rok_od', '').strip()
    rok_do = request.GET.get('rok_do', '').strip()
    if query:
        osoby = _szukaj_fts(osoby, query)
    if sektor_id:
        osoby = osoby.filter(grob__sektor_id=sektor_id)
    if typ:
        osoby = osoby.filter(grob__typ=typ)
    if rok_od.isdigit():
        osoby = osoby.filter(data_smierci__year__gte=int(rok_od))
    if rok_do.isdigit():
        osoby = osoby.filter(data_smierci__year__lte=int(rok_do))
    return osoby.order_by('nazwisko', 'imie')


@cache_page(60 * 30)
def indeks_nazwisk(request):
    nazwiska = (
        Osoba.objects.values('nazwisko')
        .annotate(n=Count('id'))
        .order_by('nazwisko')
    )
    grupy = defaultdict(list)
    for n in nazwiska:
        litera = (n['nazwisko'] or '?')[0].upper()
        if not litera.isalpha():
            litera = '#'
        grupy[litera].append({'nazwisko': n['nazwisko'], 'liczba': n['n']})
    grupy_lista = sorted(grupy.items())
    return render(request, 'groby/indeks_nazwisk.html', {
        'grupy': grupy_lista,
        'litery': [g[0] for g in grupy_lista],
        'razem': sum(n['n'] for n in nazwiska),
    })


def eksport_gedcom(request):
    """Eksport wszystkich osób + relacji do GEDCOM 5.5.5."""
    osoby = list(Osoba.objects.select_related('grob__sektor').all())
    relacje = list(Relacja.objects.select_related('osoba_a', 'osoba_b').all())

    # Pogrupuj relacje rodzic/małżeństwo na rodziny GEDCOM (FAM).
    # Dla każdej pary rodziców, wszystkie ich dzieci tworzą jedną FAM.
    # Małżeństwa bez wspólnych dzieci dostają własny FAM.
    rodzice_dziecka = defaultdict(list)  # dziecko_id -> [rodzice_ids]
    pary_malzonkow = set()
    for r in relacje:
        if r.typ == 'rodzic':
            rodzice_dziecka[r.osoba_b_id].append(r.osoba_a_id)
        elif r.typ == 'malzenstwo':
            para = tuple(sorted([r.osoba_a_id, r.osoba_b_id]))
            pary_malzonkow.add(para)

    rodziny = {}  # klucz: para rodziców (sorted) -> {ojciec, matka, dzieci}
    for dziecko_id, rodzice in rodzice_dziecka.items():
        klucz = tuple(sorted(rodzice))
        if klucz not in rodziny:
            rodziny[klucz] = {'rodzice': set(rodzice), 'dzieci': set()}
        rodziny[klucz]['dzieci'].add(dziecko_id)
    for para in pary_malzonkow:
        if para not in rodziny:
            rodziny[para] = {'rodzice': set(para), 'dzieci': set()}

    rodzina_indeks = {}  # klucz -> numer FAM
    for i, klucz in enumerate(rodziny.keys(), start=1):
        rodzina_indeks[klucz] = i

    # Mapowanie osoba -> FAMC (rodzina-z-rodzicow), FAMS (rodziny-malzeńskie)
    famc = defaultdict(list)
    fams = defaultdict(list)
    for klucz, dane in rodziny.items():
        idx = rodzina_indeks[klucz]
        for r in dane['rodzice']:
            fams[r].append(idx)
        for d in dane['dzieci']:
            famc[d].append(idx)

    linie = ['0 HEAD', '1 SOUR Informator-Cmentarny-Szydlow', '2 VERS 1.0',
             '1 GEDC', '2 VERS 5.5.5', '2 FORM LINEAGE-LINKED', '1 CHAR UTF-8']

    def gd_data(d):
        if not d:
            return None
        return d.strftime('%d %b %Y').upper()

    for o in osoby:
        linie.append(f'0 @I{o.pk}@ INDI')
        nazw = o.nazwisko_rodowe or o.nazwisko
        linie.append(f'1 NAME {o.imie} {o.drugie_imie} /{nazw}/'.replace('  ', ' ').strip())
        if o.imie:
            linie.append(f'2 GIVN {o.imie}{(" " + o.drugie_imie) if o.drugie_imie else ""}')
        if nazw:
            linie.append(f'2 SURN {nazw}')
        if o.data_urodzenia:
            linie.append('1 BIRT')
            linie.append(f'2 DATE {gd_data(o.data_urodzenia)}')
            if o.miejsce_urodzenia:
                linie.append(f'2 PLAC {o.miejsce_urodzenia}')
        if o.data_smierci:
            linie.append('1 DEAT')
            linie.append(f'2 DATE {gd_data(o.data_smierci)}')
            linie.append(f'2 PLAC Szydlow, sektor {o.grob.sektor.nazwa}, grob {o.grob.numer}')
        if o.biogram:
            linie.append(f'1 NOTE {o.biogram[:200]}')
        for f in famc[o.pk]:
            linie.append(f'1 FAMC @F{f}@')
        for f in fams[o.pk]:
            linie.append(f'1 FAMS @F{f}@')

    for klucz, dane in rodziny.items():
        idx = rodzina_indeks[klucz]
        linie.append(f'0 @F{idx}@ FAM')
        rodzice = sorted(dane['rodzice'])
        # Pierwsza osoba jako HUSB, druga jako WIFE (uproszczenie — bez płci w bazie)
        if rodzice:
            linie.append(f'1 HUSB @I{rodzice[0]}@')
        if len(rodzice) > 1:
            linie.append(f'1 WIFE @I{rodzice[1]}@')
        for d in sorted(dane['dzieci']):
            linie.append(f'1 CHIL @I{d}@')

    linie.append('0 TRLR')

    resp = HttpResponse('\r\n'.join(linie), content_type='application/x-gedcom; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="cmentarz-szydlow.ged"'
    return resp


def galeria_cmentarza(request):
    sektor_id = request.GET.get('sektor', '').strip()
    qs = Zdjecie.objects.select_related('grob__sektor').order_by('-data_dodania')
    if sektor_id:
        qs = qs.filter(grob__sektor_id=sektor_id)
    page = Paginator(qs, 60).get_page(request.GET.get('page'))
    return render(request, 'groby/galeria.html', {
        'page': page,
        'sektory': Sektor.objects.order_by('nazwa'),
        'sektor_id': sektor_id,
        'razem': qs.count(),
    })


@login_required
@require_POST
def zapisz_szukanie(request):
    nazwa = (request.POST.get('nazwa') or '').strip()[:100]
    querystring = (request.POST.get('querystring') or '').strip()[:500]
    if not nazwa or not querystring:
        messages.error(request, 'Wpisz nazwę.')
        return redirect(request.META.get('HTTP_REFERER', '/'))
    ZapisaneSzukanie.objects.create(user=request.user, nazwa=nazwa, querystring=querystring)
    messages.success(request, f'Zapisano: „{nazwa}"')
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@require_POST
def usun_zapisane(request, pk):
    ZapisaneSzukanie.objects.filter(pk=pk, user=request.user).delete()
    return redirect('groby:profil')


def lista_wpisow(request):
    typ = request.GET.get('typ', '').strip()
    qs = Wpis.objects.filter(opublikowany=True).select_related('osoba__grob__sektor', 'autor')
    if typ:
        qs = qs.filter(typ=typ)
    return render(request, 'groby/wpisy_lista.html', {
        'wpisy': qs,
        'typ': typ,
        'typy': Wpis.TYP_CHOICES,
    })


def wpis_detail(request, slug):
    wpis = get_object_or_404(Wpis, slug=slug, opublikowany=True)
    return render(request, 'groby/wpis_detail.html', {'wpis': wpis})


def timeline(request):
    """Oś czasu: paski życia osób na linii czasu."""
    tag_slug = request.GET.get('tag', '').strip()
    qs = Osoba.objects.exclude(data_smierci__isnull=True).select_related('grob__sektor')
    if tag_slug:
        qs = qs.filter(tagi__slug=tag_slug)
    qs = qs.order_by('data_urodzenia', 'data_smierci')[:1000]
    osoby = []
    for o in qs:
        urodzony = o.data_urodzenia.year if o.data_urodzenia else (o.data_smierci.year - 70)
        zmarl = o.data_smierci.year
        osoby.append({
            'pk': o.pk, 'nazwa': f'{o.imie} {o.nazwisko}',
            'od': urodzony, 'do': zmarl, 'wiek': zmarl - urodzony,
            'sektor': o.grob.sektor.nazwa, 'numer': o.grob.numer,
        })
    return render(request, 'groby/timeline.html', {
        'dane_json': json.dumps(osoby, ensure_ascii=False),
        'wszystkie_tagi': Tag.objects.all(),
        'aktywny_tag': tag_slug,
    })


def widget(request, typ='szukaj'):
    """Uproszczone widoki bez nawigacji do osadzania w iframe innych stron."""
    osoby = []
    query = (request.GET.get('q') or '').strip()
    if query:
        osoby = list(_szukaj_fts(Osoba.objects.select_related('grob__sektor'), query)[:30])
    return render(request, 'groby/widget.html', {
        'osoby': osoby,
        'query': query,
    })


def health_check(request):
    """Endpoint dla monitoringu — zwraca JSON ze stanem systemu."""
    from django.db import connection
    import shutil
    stan = {'status': 'ok', 'checks': {}}
    try:
        with connection.cursor() as c:
            c.execute('SELECT 1')
            c.fetchone()
        stan['checks']['database'] = 'ok'
    except Exception as e:
        stan['status'] = 'fail'
        stan['checks']['database'] = f'fail: {e}'
    try:
        usage = shutil.disk_usage('/')
        stan['checks']['disk_free_gb'] = round(usage.free / (1024**3), 2)
        if usage.free < 100 * 1024**2:  # < 100 MB
            stan['status'] = 'warn'
            stan['checks']['disk'] = 'warn: low space'
    except Exception:
        pass
    stan['liczby'] = {
        'sektory': Sektor.objects.count(),
        'groby': Grob.objects.count(),
        'osoby': Osoba.objects.count(),
        'zdjecia': Zdjecie.objects.count(),
    }
    return JsonResponse(stan, status=200 if stan['status'] == 'ok' else 503)


def szukaj_na_mapie(request):
    """API zwracające pasujące groby z pozycją na planie — dla wyszukiwarki na mapie."""
    q = (request.GET.get('q') or '').strip()
    if len(q) < 2:
        return JsonResponse({'wyniki': []})
    osoby_ids = list(_szukaj_fts(Osoba.objects.values_list('pk', flat=True), q)[:50])
    groby = Grob.objects.filter(
        Q(osoby__pk__in=osoby_ids) | Q(numer__icontains=q),
        plan_x__isnull=False, plan_y__isnull=False,
    ).select_related('sektor').prefetch_related('osoby').distinct()[:30]
    wyniki = []
    for g in groby:
        osoby_str = ', '.join(f'{o.imie} {o.nazwisko}' for o in g.osoby.all()[:3])
        wyniki.append({
            'id': g.pk, 'x': g.plan_x, 'y': g.plan_y,
            'sektor': g.sektor.nazwa, 'numer': g.numer, 'rzad': g.rzad,
            'osoby': osoby_str or '—',
            'url': reverse('groby:grob_detail', args=[g.pk]),
        })
    return JsonResponse({'wyniki': wyniki})


def lista_panoram(request):
    panoramy = Panorama.objects.select_related('sektor').order_by('kolejnosc', 'nazwa')
    return render(request, 'groby/panoramy_lista.html', {'panoramy': panoramy})


def panorama_detail(request, pk):
    panorama = get_object_or_404(Panorama.objects.prefetch_related('hotspoty__grob__sektor'), pk=pk)
    return render(request, 'groby/panorama_detail.html', {'panorama': panorama})


# ---- Magic link ----

def magic_link_zarzadaj(request):
    if request.method != 'POST':
        return render(request, 'groby/magic_link.html')
    email = (request.POST.get('email') or '').strip()
    if not email:
        messages.error(request, 'Podaj adres e-mail.')
        return redirect('groby:magic_link')
    User = __import__('django.contrib.auth', fromlist=['get_user_model']).get_user_model()
    user = User.objects.filter(email__iexact=email).first()
    if user:
        import secrets
        from datetime import timedelta
        from django.utils import timezone
        from django.core.mail import send_mail
        from django.conf import settings
        token = secrets.token_urlsafe(32)
        TokenLogowania.objects.create(
            user=user, token=token,
            data_wygasniecia=timezone.now() + timedelta(minutes=30),
        )
        link = request.build_absolute_uri(reverse('groby:magic_link_uzyj', args=[token]))
        send_mail(
            'Logowanie do Informatora Cmentarnego',
            f'Kliknij aby się zalogować: {link}\nLink wygaśnie za 30 minut.\n',
            settings.DEFAULT_FROM_EMAIL, [email], fail_silently=True,
        )
    messages.success(request, 'Jeśli ten e-mail jest zarejestrowany, link został wysłany.')
    return redirect('groby:magic_link')


def magic_link_uzyj(request, token):
    from django.utils import timezone
    t = TokenLogowania.objects.filter(token=token, wykorzystany=False).first()
    if not t or t.data_wygasniecia < timezone.now():
        messages.error(request, 'Link nieaktualny lub już użyty. Wygeneruj nowy.')
        return redirect('groby:magic_link')
    t.wykorzystany = True
    t.save(update_fields=['wykorzystany'])
    login(request, t.user, backend='django.contrib.auth.backends.ModelBackend')
    messages.success(request, f'Zalogowano jako {t.user.username}.')
    return redirect('groby:profil')


# ---- 2FA TOTP ----

@login_required
def totp_setup(request):
    """Konfiguracja 2FA: generuje sekret, pokazuje QR; potwierdzenie kodem aktywuje urządzenie."""
    from django_otp.plugins.otp_totp.models import TOTPDevice
    import qrcode, base64
    device = TOTPDevice.objects.filter(user=request.user, name='default').first()

    if request.method == 'POST':
        kod = (request.POST.get('kod') or '').strip()
        if device and device.verify_token(kod):
            device.confirmed = True
            device.save()
            messages.success(request, '2FA aktywne.')
            return redirect('groby:profil')
        messages.error(request, 'Kod nieprawidłowy.')

    if not device:
        device = TOTPDevice.objects.create(user=request.user, name='default', confirmed=False)
    config_url = device.config_url
    img = qrcode.make(config_url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()
    return render(request, 'groby/totp_setup.html', {
        'device': device, 'qr_b64': qr_b64, 'config_url': config_url,
    })


@login_required
@require_POST
def totp_wylacz(request):
    from django_otp.plugins.otp_totp.models import TOTPDevice
    TOTPDevice.objects.filter(user=request.user).delete()
    messages.success(request, '2FA wyłączone.')
    return redirect('groby:profil')


# ---- Web Push ----

def push_klucz_publiczny(request):
    """Zwraca klucz publiczny VAPID (frontend potrzebuje do subscribe())."""
    from django.conf import settings
    return JsonResponse({'publicKey': getattr(settings, 'VAPID_PUBLIC_KEY', '')})


@require_POST
def push_subskrybuj(request):
    try:
        data = json.loads(request.body or '{}')
        endpoint = data['endpoint']
        keys = data.get('keys', {})
        SubskrypcjaPush.objects.update_or_create(
            endpoint=endpoint,
            defaults={
                'p256dh': keys.get('p256dh', '')[:200],
                'auth': keys.get('auth', '')[:80],
                'user': request.user if request.user.is_authenticated else None,
                'user_agent': request.META.get('HTTP_USER_AGENT', '')[:255],
            },
        )
    except (KeyError, json.JSONDecodeError):
        return JsonResponse({'ok': False}, status=400)
    return JsonResponse({'ok': True})


# ---- Bulk import zdjęć ZIP ----

@login_required
def bulk_import_zdjec(request):
    if not request.user.is_staff:
        return redirect('admin:login')
    raport = None
    if request.method == 'POST' and request.FILES.get('zip'):
        import zipfile, re, os
        from django.core.files.base import ContentFile
        plik = request.FILES['zip']
        dodanych, pominietych, bledow = 0, 0, []
        try:
            with zipfile.ZipFile(plik) as zf:
                for nazwa in zf.namelist():
                    if nazwa.endswith('/') or not re.search(r'\.(jpe?g|png|webp)$', nazwa, re.I):
                        continue
                    base = os.path.basename(nazwa)
                    # Wzorzec: SEKTOR_NUMER[_KOL].ext  np. A_12.jpg, B_5_2.jpg
                    m = re.match(r'^([A-Za-z0-9]+)[_-]([A-Za-z0-9]+)(?:[_-](\d+))?\..+$', base)
                    if not m:
                        bledow.append(f'{base}: nie pasuje wzorzec SEKTOR_NUMER.jpg')
                        continue
                    sektor_n, numer, kol = m.group(1), m.group(2), int(m.group(3) or 0)
                    grob = Grob.objects.filter(sektor__nazwa__iexact=sektor_n, numer=numer).first()
                    if not grob:
                        pominietych += 1
                        bledow.append(f'{base}: brak grobu {sektor_n}/{numer}')
                        continue
                    z = Zdjecie(grob=grob, kolejnosc=kol)
                    z.plik.save(base, ContentFile(zf.read(nazwa)), save=True)
                    dodanych += 1
            raport = {'dodanych': dodanych, 'pominietych': pominietych, 'bledow': bledow[:30]}
        except zipfile.BadZipFile:
            messages.error(request, 'Nieprawidłowy plik ZIP.')
    return render(request, 'groby/bulk_import.html', {'raport': raport})


def wspolni_przodkowie(request):
    """Znajduje wspólnych przodków dwóch osób przez analizę grafu Relacja typ='rodzic'."""
    a_id = request.GET.get('a', '').strip()
    b_id = request.GET.get('b', '').strip()
    osoba_a = osoba_b = None
    wspolni = []
    if a_id.isdigit() and b_id.isdigit():
        osoba_a = Osoba.objects.filter(pk=a_id).select_related('grob__sektor').first()
        osoba_b = Osoba.objects.filter(pk=b_id).select_related('grob__sektor').first()
        if osoba_a and osoba_b:
            przodkowie_a = _zbierz_przodkow(osoba_a.pk)
            przodkowie_b = _zbierz_przodkow(osoba_b.pk)
            wspolni_ids = set(przodkowie_a) & set(przodkowie_b)
            wspolni = list(Osoba.objects.filter(pk__in=wspolni_ids).select_related('grob__sektor'))
    return render(request, 'groby/wspolni_przodkowie.html', {
        'osoba_a': osoba_a, 'osoba_b': osoba_b, 'wspolni': wspolni,
    })


def _zbierz_przodkow(osoba_id, glebokosc=10):
    """BFS w górę w grafie rodzic→dziecko (rodzice osoby x's są A w Relacji rodzic z B=x)."""
    from collections import deque
    odwiedzone = set()
    kolejka = deque([(osoba_id, 0)])
    while kolejka:
        bieg, d = kolejka.popleft()
        if d >= glebokosc:
            continue
        rodzice = Relacja.objects.filter(typ='rodzic', osoba_b_id=bieg).values_list('osoba_a_id', flat=True)
        for r in rodzice:
            if r not in odwiedzone:
                odwiedzone.add(r)
                kolejka.append((r, d + 1))
    return odwiedzone


@require_POST
def dodaj_komentarz(request, wspomnienie_id):
    if _antybot(request):
        return redirect(request.META.get('HTTP_REFERER', '/'))
    w = get_object_or_404(Wspomnienie, pk=wspomnienie_id, status='zaakceptowane')
    tresc = (request.POST.get('tresc') or '').strip()
    if not tresc:
        messages.error(request, 'Treść komentarza pusta.')
        return redirect('groby:osoba_detail', pk=w.osoba_id)
    parent_id = request.POST.get('parent', '').strip()
    parent = Komentarz.objects.filter(pk=parent_id, wspomnienie=w).first() if parent_id else None
    Komentarz.objects.create(
        wspomnienie=w, parent=parent,
        autor_user=request.user if request.user.is_authenticated else None,
        autor_imie=(request.POST.get('autor_imie') or '').strip()[:100],
        tresc=tresc[:2000],
    )
    messages.success(request, 'Komentarz dodany. Pojawi się po akceptacji.')
    return redirect('groby:osoba_detail', pk=w.osoba_id)


def karta_grobu_pdf(request, pk):
    """Pełen PDF z kartą grobu: foto, dane, osoby, QR, lokalizacja."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
    from pathlib import Path
    import qrcode

    grob = get_object_or_404(Grob.objects.select_related('sektor').prefetch_related('osoby', 'zdjecia'), pk=pk)

    czcionka = 'Helvetica'
    for k in (r'C:\Windows\Fonts\arial.ttf',):
        if Path(k).exists():
            try:
                pdfmetrics.registerFont(TTFont('Polska', k))
                czcionka = 'Polska'
                break
            except Exception:
                pass

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    sty = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', fontName=czcionka, fontSize=22, leading=26, spaceAfter=8, textColor=colors.HexColor('#2e4430'))
    h2 = ParagraphStyle('h2', fontName=czcionka, fontSize=14, leading=18, spaceAfter=4, textColor=colors.HexColor('#3a553a'))
    body = ParagraphStyle('b', fontName=czcionka, fontSize=10, leading=14)
    meta = ParagraphStyle('m', fontName=czcionka, fontSize=9, leading=12, textColor=colors.HexColor('#4c6b4a'))

    elementy = []
    elementy.append(Paragraph('Karta grobu', meta))
    tytul = f'Sektor {grob.sektor.nazwa}{" / Rząd " + grob.rzad if grob.rzad else ""} / Grób nr {grob.numer}'
    elementy.append(Paragraph(tytul, h1))
    elementy.append(Paragraph(f'Typ: {grob.get_typ_display()}', body))
    elementy.append(Spacer(1, 0.5*cm))

    # QR po prawej, foto po lewej (Tabela)
    url = request.build_absolute_uri(reverse('groby:grob_detail', args=[grob.pk]))
    qr_img = qrcode.make(url)
    qr_buf = io.BytesIO()
    qr_img.save(qr_buf, format='PNG')
    qr_buf.seek(0)

    foto = None
    if grob.zdjecia.exists():
        z = grob.zdjecia.first()
        try:
            foto = Image(z.plik.path, width=8*cm, height=8*cm, kind='proportional')
        except Exception:
            foto = None
    qr = Image(qr_buf, width=4*cm, height=4*cm)

    if foto:
        tab = Table([[foto, qr]], colWidths=[10*cm, 4*cm])
    else:
        tab = Table([[Paragraph('Brak zdjęcia', meta), qr]], colWidths=[10*cm, 4*cm])
    tab.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elementy.append(tab)
    elementy.append(Spacer(1, 0.5*cm))

    elementy.append(Paragraph('Osoby pochowane:', h2))
    if grob.osoby.exists():
        for o in grob.osoby.all():
            ur = o.data_urodzenia.strftime('%d.%m.%Y') if o.data_urodzenia else '?'
            sm = o.data_smierci.strftime('%d.%m.%Y') if o.data_smierci else '?'
            tekst = f'<b>{o.nazwisko} {o.imie}</b>'
            if o.nazwisko_rodowe:
                tekst += f' (z d. {o.nazwisko_rodowe})'
            tekst += f' — ur. {ur}, zm. {sm}'
            if o.wiek:
                tekst += f', żył(a) {o.wiek} lat'
            elementy.append(Paragraph(tekst, body))
            if o.miejsce_urodzenia:
                elementy.append(Paragraph(f'  Miejsce ur.: {o.miejsce_urodzenia}', meta))
            elementy.append(Spacer(1, 0.2*cm))
    else:
        elementy.append(Paragraph('Brak zarejestrowanych osób.', meta))

    elementy.append(Spacer(1, 0.5*cm))
    elementy.append(Paragraph('Cmentarz parafialny w Szydłowie · ' + url, meta))

    doc.build(elementy)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="grob-{grob.sektor.nazwa}-{grob.numer}.pdf"'
    return resp


def kronika(request):
    """Publiczna kronika ostatnich zmian (zdjęć, biogramów, wspomnień)."""
    zmiany = HistoriaZmian.objects.select_related('user').order_by('-data')[:50]
    nowe_zdjecia = Zdjecie.objects.select_related('grob__sektor').order_by('-data_dodania')[:20]
    nowe_wspomnienia = Wspomnienie.objects.filter(status='zaakceptowane').select_related('osoba__grob__sektor').order_by('-data_dodania')[:20]
    return render(request, 'groby/kronika.html', {
        'zmiany': zmiany,
        'nowe_zdjecia': nowe_zdjecia,
        'nowe_wspomnienia': nowe_wspomnienia,
    })


def lista_tras(request):
    trasy = Trasa.objects.filter(opublikowana=True).prefetch_related('punkty__grob__sektor')
    return render(request, 'groby/trasy_lista.html', {'trasy': trasy})


def trasa_detail(request, slug):
    trasa = get_object_or_404(Trasa, slug=slug, opublikowana=True)
    punkty = trasa.punkty.select_related('grob__sektor').prefetch_related('grob__osoby')
    dane = [{
        'kolejnosc': p.kolejnosc, 'podpis': p.podpis,
        'grob_id': p.grob.pk, 'sektor': p.grob.sektor.nazwa, 'numer': p.grob.numer,
        'plan_x': p.grob.plan_x, 'plan_y': p.grob.plan_y,
        'osoby': ', '.join(f'{o.imie} {o.nazwisko}' for o in p.grob.osoby.all()[:3]),
    } for p in punkty]
    return render(request, 'groby/trasa_detail.html', {
        'trasa': trasa, 'punkty': punkty, 'dane_json': json.dumps(dane, ensure_ascii=False),
    })


def skaner_qr(request):
    return render(request, 'groby/skaner.html')


def newsletter_zapis(request):
    if request.method == 'POST':
        if _antybot(request):
            return redirect(request.META.get('HTTP_REFERER', '/'))
        email = (request.POST.get('email') or '').strip().lower()[:150]
        if email:
            import secrets
            token = secrets.token_urlsafe(32)
            Newsletter.objects.update_or_create(
                email=email,
                defaults={'aktywny': True, 'token_anulowania': token},
            )
            messages.success(request, f'Zapisano {email} do newslettera.')
        else:
            messages.error(request, 'Podaj e-mail.')
    return redirect(request.META.get('HTTP_REFERER', '/'))


def newsletter_anuluj(request, token):
    n = Newsletter.objects.filter(token_anulowania=token).first()
    if n:
        n.aktywny = False
        n.save(update_fields=['aktywny'])
        messages.success(request, f'Anulowano subskrypcję {n.email}.')
    return redirect('groby:home')


# ---- Kwiaty ----

@require_POST
def zloz_kwiat(request, pk):
    if _antybot(request):
        return redirect('groby:osoba_detail', pk=pk)
    osoba = get_object_or_404(Osoba, pk=pk)
    from datetime import timedelta
    from django.utils import timezone
    iph = _hash_ip(request)
    if Kwiat.objects.filter(osoba=osoba, ip_hash=iph,
                            data_zlozenia__gte=timezone.now() - timedelta(minutes=5)).exists():
        return redirect('groby:osoba_detail', pk=pk)
    Kwiat.objects.create(
        osoba=osoba,
        autor_user=request.user if request.user.is_authenticated else None,
        autor_imie=(request.POST.get('autor_imie') or '').strip()[:100],
        wiadomosc=(request.POST.get('wiadomosc') or '').strip()[:300],
        rodzaj=(request.POST.get('rodzaj') or 'roza')[:20],
        ip_hash=iph,
    )
    messages.success(request, '💐 Kwiat złożony — pozostanie 7 dni.')
    return redirect('groby:osoba_detail', pk=pk)


# ---- Memory wall ----

def memory_wall(request):
    qs = Wspomnienie.objects.filter(status='zaakceptowane').select_related('osoba__grob__sektor', 'autor_user').order_by('-data_dodania')
    page = Paginator(qs, 30).get_page(request.GET.get('page'))
    return render(request, 'groby/memory_wall.html', {'page': page})


# ---- Kto żył w roku X ----

def kto_zyl(request):
    rok = request.GET.get('rok', '').strip()
    osoby = []
    if rok.isdigit():
        r = int(rok)
        # Żyjący w danym roku: data_urodzenia <= rok <= data_smierci
        qs = Osoba.objects.filter(
            data_urodzenia__year__lte=r,
            data_smierci__year__gte=r,
        ).select_related('grob__sektor')
        osoby = list(qs[:500])
    return render(request, 'groby/kto_zyl.html', {
        'rok': rok, 'osoby': osoby,
    })


# ---- Mapa pochodzenia (geo) ----

def mapa_pochodzenia(request):
    """Geo-mapa miejsc urodzenia. Geokodowanie z cache w GeoCache."""
    miejsca = (Osoba.objects
               .exclude(miejsce_urodzenia='')
               .values('miejsce_urodzenia')
               .annotate(c=Count('id'))
               .order_by('-c')[:200])
    dane = []
    for m in miejsca:
        nazwa = m['miejsce_urodzenia']
        cache = GeoCache.objects.filter(nazwa=nazwa).first()
        if cache and cache.lat is not None:
            dane.append({'nazwa': nazwa, 'liczba': m['c'], 'lat': cache.lat, 'lng': cache.lng})
    return render(request, 'groby/mapa_pochodzenia.html', {
        'dane_json': json.dumps(dane, ensure_ascii=False),
        'razem_miejsc': len(miejsca),
        'zgeokodowane': len(dane),
    })


# ---- Moja rodzina ----

@login_required
def moja_rodzina(request):
    profil_obj, _ = Profil.objects.get_or_create(user=request.user)
    obs_osoby = list(profil_obj.obserwowane_osoby.all())
    obs_groby_osoby = []
    for g in profil_obj.obserwowane_groby.prefetch_related('osoby'):
        obs_groby_osoby += list(g.osoby.all())
    wszystkie = list({o.pk: o for o in obs_osoby + obs_groby_osoby}.values())
    relacje = []
    if wszystkie:
        ids = [o.pk for o in wszystkie]
        relacje = list(Relacja.objects.filter(
            Q(osoba_a__in=ids) | Q(osoba_b__in=ids),
        ).select_related('osoba_a', 'osoba_b'))
    return render(request, 'groby/moja_rodzina.html', {
        'osoby': wszystkie, 'relacje': relacje,
    })


# ---- Inline edit (staff) ----

@login_required
@require_POST
def inline_edit(request, model, pk, pole):
    if not request.user.is_staff:
        return JsonResponse({'ok': False, 'error': 'Brak uprawnień'}, status=403)
    DOZWOLONE = {'Grob': ('uwagi', 'numer_aktu', 'rodzaj_opis'),
                 'Osoba': ('biogram', 'miejsce_urodzenia')}
    if model not in DOZWOLONE or pole not in DOZWOLONE[model]:
        return JsonResponse({'ok': False, 'error': 'Pole niedozwolone'}, status=400)
    Klasa = Grob if model == 'Grob' else Osoba
    obj = get_object_or_404(Klasa, pk=pk)
    nowa_wartosc = (request.POST.get('wartosc') or '').strip()
    setattr(obj, pole, nowa_wartosc)
    obj.save(update_fields=[pole, 'data_modyfikacji'] if hasattr(obj, 'data_modyfikacji') else [pole])
    return JsonResponse({'ok': True, 'wartosc': nowa_wartosc})


# ---- Drag & drop reorder zdjęć ----

@login_required
@require_POST
def przestaw_zdjecia(request):
    if not request.user.is_staff:
        return JsonResponse({'ok': False}, status=403)
    try:
        kolejnosc = json.loads(request.body)['kolejnosc']
    except (KeyError, json.JSONDecodeError):
        return JsonResponse({'ok': False}, status=400)
    for idx, z_id in enumerate(kolejnosc):
        Zdjecie.objects.filter(pk=z_id).update(kolejnosc=idx)
    return JsonResponse({'ok': True})


# ---- Plebiscyt nagrobków ----

def plebiscyt(request):
    okres_dni = int(request.GET.get('dni', '30'))
    from datetime import timedelta
    from django.utils import timezone
    granica = timezone.now() - timedelta(days=okres_dni)
    ranking = (Grob.objects
               .annotate(liczba_glosow=Count('glosy', filter=Q(glosy__data__gte=granica)))
               .filter(liczba_glosow__gt=0)
               .select_related('sektor')
               .prefetch_related('zdjecia', 'osoby')
               .order_by('-liczba_glosow')[:30])
    return render(request, 'groby/plebiscyt.html', {
        'ranking': ranking, 'dni': okres_dni,
    })


@require_POST
def glosuj(request, pk):
    if _antybot(request):
        return redirect(request.META.get('HTTP_REFERER', '/'))
    grob = get_object_or_404(Grob, pk=pk)
    iph = _hash_ip(request)
    _, created = GlosNagrobek.objects.get_or_create(
        grob=grob, ip_hash=iph,
        defaults={'user': request.user if request.user.is_authenticated else None},
    )
    if created:
        messages.success(request, '✓ Głos oddany!')
    else:
        messages.info(request, 'Już głosowałeś na ten grób.')
    return redirect(request.META.get('HTTP_REFERER', '/'))


# ---- Statystyki użytkownika ----

@login_required
def moje_statystyki(request):
    user = request.user
    return render(request, 'groby/moje_statystyki.html', {
        'liczba_swiec': user.swiece.count(),
        'liczba_kwiatow': user.kwiat_set.count() if hasattr(user, 'kwiat_set') else Kwiat.objects.filter(autor_user=user).count(),
        'liczba_wspomnien': user.wspomnienia.count(),
        'liczba_komentarzy': Komentarz.objects.filter(autor_user=user).count(),
        'liczba_zgloszen': user.zgloszenia.count(),
        'liczba_obserwowanych_grobow': user.profil.obserwowane_groby.count() if hasattr(user, 'profil') else 0,
        'liczba_obserwowanych_osob': user.profil.obserwowane_osoby.count() if hasattr(user, 'profil') else 0,
        'liczba_zapisanych_szukan': user.zapisane_szukania.count(),
        'liczba_glosow': GlosNagrobek.objects.filter(user=user).count(),
        'liczba_odznak': user.odznaki.count(),
        'odznaki': user.odznaki.select_related('odznaka'),
    })


# ---- Intencje mszalne ----

def intencje_form(request):
    if request.method == 'POST':
        if _antybot(request):
            return redirect('groby:intencje_form')
        intencja = (request.POST.get('intencja') or '').strip()
        imie = (request.POST.get('imie') or '').strip()[:200]
        email = (request.POST.get('email') or '').strip()[:200]
        if not (intencja and imie and email):
            messages.error(request, 'Wymagane: imię, e-mail, intencja.')
            return redirect('groby:intencje_form')
        osoba_id = request.POST.get('osoba_id', '').strip()
        osoba = Osoba.objects.filter(pk=osoba_id).first() if osoba_id.isdigit() else None
        proponowana = request.POST.get('data') or None
        IntencjaMszalna.objects.create(
            osoba=osoba,
            zamawiajacy_imie=imie,
            zamawiajacy_email=email,
            zamawiajacy_tel=(request.POST.get('tel') or '').strip()[:30],
            intencja=intencja[:2000],
            proponowana_data=proponowana if proponowana else None,
        )
        messages.success(request, 'Intencja przyjęta. Skontaktujemy się drogą e-mail.')
        return redirect('groby:intencje_form')
    return render(request, 'groby/intencje.html')


# ---- Zaproszenia do edycji ----

@login_required
def utworz_zaproszenie(request, osoba_id):
    if not request.user.is_authenticated:
        return redirect('admin:login')
    osoba = get_object_or_404(Osoba, pk=osoba_id)
    if request.method != 'POST':
        return redirect('groby:osoba_detail', pk=osoba_id)
    email = (request.POST.get('email') or '').strip()
    if not email:
        return redirect('groby:osoba_detail', pk=osoba_id)
    import secrets
    from django.core.mail import send_mail
    from django.conf import settings
    token = secrets.token_urlsafe(32)
    Zaproszenie.objects.create(osoba=osoba, email=email, token=token, autor=request.user)
    link = request.build_absolute_uri(reverse('groby:przyjmij_zaproszenie', args=[token]))
    send_mail(
        f'Zaproszenie do współedycji — {osoba}',
        f'Otrzymałeś/aś zaproszenie do współedycji wpisu o osobie {osoba}.\n'
        f'Kliknij aby skorzystać: {link}\n',
        settings.DEFAULT_FROM_EMAIL, [email], fail_silently=True,
    )
    messages.success(request, f'Zaproszenie wysłane na {email}.')
    return redirect('groby:osoba_detail', pk=osoba_id)


@login_required
def przyjmij_zaproszenie(request, token):
    from django.utils import timezone
    z = get_object_or_404(Zaproszenie, token=token, wykorzystane_przez__isnull=True)
    z.wykorzystane_przez = request.user
    z.data_wykorzystania = timezone.now()
    z.save()
    messages.success(request, f'Możesz teraz edytować {z.osoba} w panelu admin.')
    return redirect('groby:osoba_detail', pk=z.osoba.pk)


# ---- GDPR ----

def polityka_prywatnosci(request):
    return render(request, 'groby/polityka.html')


def regulamin(request):
    return render(request, 'groby/regulamin.html')


@login_required
def eksport_moich_danych(request):
    """GDPR Right to Access — ZIP z wszystkimi danymi użytkownika."""
    import zipfile
    user = request.user
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        dane = {
            'user': {
                'username': user.username, 'email': user.email,
                'first_name': user.first_name, 'last_name': user.last_name,
                'date_joined': user.date_joined.isoformat(),
            },
            'profil': None,
            'obserwowane_groby': [],
            'obserwowane_osoby': [],
            'wspomnienia': [],
            'komentarze': [],
            'swieczki': [],
            'kwiaty': [],
            'zgloszenia': [],
            'zapisane_szukania': [],
            'odznaki': [],
            'glosy_plebiscytu': [],
        }
        if hasattr(user, 'profil'):
            dane['profil'] = {
                'pokrewienstwo': user.profil.pokrewienstwo,
                'data_utworzenia': user.profil.data_utworzenia.isoformat(),
            }
            dane['obserwowane_groby'] = [str(g) for g in user.profil.obserwowane_groby.all()]
            dane['obserwowane_osoby'] = [str(o) for o in user.profil.obserwowane_osoby.all()]
        for w in user.wspomnienia.all():
            dane['wspomnienia'].append({'osoba': str(w.osoba), 'tresc': w.tresc, 'data': w.data_dodania.isoformat()})
        for s in user.swiece.all():
            dane['swieczki'].append({'osoba': str(s.osoba), 'intencja': s.intencja, 'data': s.data_zapalenia.isoformat()})
        for k in Kwiat.objects.filter(autor_user=user):
            dane['kwiaty'].append({'osoba': str(k.osoba), 'rodzaj': k.rodzaj, 'wiadomosc': k.wiadomosc, 'data': k.data_zlozenia.isoformat()})
        for k in Komentarz.objects.filter(autor_user=user):
            dane['komentarze'].append({'wspomnienie': str(k.wspomnienie), 'tresc': k.tresc, 'data': k.data_dodania.isoformat()})
        for z in user.zgloszenia.all():
            dane['zgloszenia'].append({'tresc': z.tresc, 'status': z.get_status_display(), 'data': z.data_dodania.isoformat()})
        for s in user.zapisane_szukania.all():
            dane['zapisane_szukania'].append({'nazwa': s.nazwa, 'querystring': s.querystring})
        for o in user.odznaki.all():
            dane['odznaki'].append({'odznaka': o.odznaka.nazwa, 'data': o.data_zdobycia.isoformat()})
        for g in GlosNagrobek.objects.filter(user=user):
            dane['glosy_plebiscytu'].append({'grob': str(g.grob), 'data': g.data.isoformat()})
        zf.writestr('moje-dane.json', json.dumps(dane, ensure_ascii=False, indent=2))
        zf.writestr('README.txt',
            'Eksport Twoich danych z Informatora Cmentarnego w Szydłowie.\n'
            'Zgodnie z RODO art. 15 (prawo dostępu).\n'
            f'Wygenerowano: {__import__("django.utils.timezone", fromlist=["now"]).now().isoformat()}\n')
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="moje-dane-{user.username}.zip"'
    return resp


@login_required
def usun_konto(request):
    if request.method == 'POST':
        potwierdzenie = (request.POST.get('potwierdzenie') or '').strip()
        if potwierdzenie != 'USUN':
            messages.error(request, 'Wpisz USUN aby potwierdzić.')
            return redirect('groby:usun_konto')
        user = request.user
        from django.contrib.auth import logout
        logout(request)
        user.delete()
        messages.success(request, 'Konto i wszystkie powiązane dane zostały usunięte.')
        return redirect('groby:home')
    return render(request, 'groby/usun_konto.html')


# ---- Top 10 / rankingi ----

def top10(request):
    """Strona z 10 rankingami."""
    from django.db.models import F, ExpressionWrapper, IntegerField, Avg
    najmlodsi = list(Osoba.objects.exclude(data_urodzenia__isnull=True).exclude(data_smierci__isnull=True)
                     .annotate(wiek_lat=ExpressionWrapper(F('data_smierci__year') - F('data_urodzenia__year'), output_field=IntegerField()))
                     .filter(wiek_lat__gte=0).order_by('wiek_lat')[:10])
    najstarsi = list(Osoba.objects.exclude(data_urodzenia__isnull=True).exclude(data_smierci__isnull=True)
                     .annotate(wiek_lat=ExpressionWrapper(F('data_smierci__year') - F('data_urodzenia__year'), output_field=IntegerField()))
                     .order_by('-wiek_lat')[:10])
    najczestsze_nazwiska = list(Osoba.objects.values('nazwisko').annotate(c=Count('id')).order_by('-c')[:10])
    najwieksze_groby = list(Grob.objects.annotate(n=Count('osoby')).order_by('-n').select_related('sektor')[:10])
    najnowsze_dodane = list(Osoba.objects.select_related('grob__sektor').order_by('-pk')[:10])
    return render(request, 'groby/top10.html', {
        'najmlodsi': najmlodsi,
        'najstarsi': najstarsi,
        'najczestsze_nazwiska': najczestsze_nazwiska,
        'najwieksze_groby': najwieksze_groby,
        'najnowsze_dodane': najnowsze_dodane,
    })


# ---- Word cloud nazwisk ----

def word_cloud(request):
    nazwiska = list(Osoba.objects.values('nazwisko').annotate(c=Count('id')).order_by('-c')[:100])
    return render(request, 'groby/word_cloud.html', {
        'nazwiska_json': json.dumps(nazwiska, ensure_ascii=False),
    })


# ---- Family tree całego cmentarza ----

def family_tree_cmentarza(request):
    osoby = list(Osoba.objects.values('id', 'imie', 'nazwisko'))
    relacje = list(Relacja.objects.values('osoba_a_id', 'osoba_b_id', 'typ'))
    nodes = [{'id': o['id'], 'name': f"{o['imie']} {o['nazwisko']}"} for o in osoby]
    links = [{'source': r['osoba_a_id'], 'target': r['osoba_b_id'], 'typ': r['typ']} for r in relacje]
    return render(request, 'groby/family_tree.html', {
        'nodes_json': json.dumps(nodes, ensure_ascii=False),
        'links_json': json.dumps(links, ensure_ascii=False),
    })


# ---- Kolaż wspomnień PDF ----

def kolaz_wspomnien_pdf(request, osoba_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from pathlib import Path

    osoba = get_object_or_404(Osoba.objects.prefetch_related('wspomnienia'), pk=osoba_id)
    wsp = osoba.wspomnienia.filter(status='zaakceptowane')

    czcionka = 'Helvetica'
    for k in (r'C:\Windows\Fonts\arial.ttf',):
        if Path(k).exists():
            try:
                pdfmetrics.registerFont(TTFont('Polska', k))
                czcionka = 'Polska'
                break
            except Exception:
                pass

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    h1 = ParagraphStyle('h1', fontName=czcionka, fontSize=22, leading=28, spaceAfter=12, alignment=1)
    sub = ParagraphStyle('sub', fontName=czcionka, fontSize=11, leading=14, spaceAfter=20, alignment=1, textColor=colors.HexColor('#4c6b4a'))
    body = ParagraphStyle('b', fontName=czcionka, fontSize=11, leading=16, spaceAfter=10, leftIndent=12, rightIndent=12)
    autor = ParagraphStyle('a', fontName=czcionka, fontSize=9, leading=12, spaceAfter=20, alignment=2, textColor=colors.HexColor('#94ae93'))

    e = []
    e.append(Paragraph('Pamiętamy', sub))
    e.append(Paragraph(f'{osoba.imie} {osoba.nazwisko}', h1))
    if osoba.data_urodzenia or osoba.data_smierci:
        e.append(Paragraph(f"{osoba.data_urodzenia.strftime('%d.%m.%Y') if osoba.data_urodzenia else '?'} — {osoba.data_smierci.strftime('%d.%m.%Y') if osoba.data_smierci else '?'}", sub))
    e.append(Spacer(1, 1*cm))
    if wsp:
        for w in wsp:
            e.append(Paragraph(f'„{w.tresc}"', body))
            e.append(Paragraph(f'— {w.autor_str}, {w.data_dodania:%d.%m.%Y}', autor))
    else:
        e.append(Paragraph('Brak wspomnień. Dodaj swoje na stronie cmentarza.', body))

    doc.build(e)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="wspomnienia-{osoba.nazwisko}.pdf"'
    return resp


# ---- Eksport galerii sektora ZIP ----

@login_required
def eksport_galerii_sektora(request, sektor_id):
    if not request.user.is_staff:
        return redirect('admin:login')
    import zipfile
    sektor = get_object_or_404(Sektor, pk=sektor_id)
    zdjecia = Zdjecie.objects.filter(grob__sektor=sektor).select_related('grob')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for z in zdjecia:
            try:
                with open(z.plik.path, 'rb') as f:
                    nazwa = f'{z.grob.numer}_{z.pk}.jpg'
                    zf.writestr(nazwa, f.read())
            except (OSError, ValueError):
                continue
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/zip')
    resp['Content-Disposition'] = f'attachment; filename="galeria-sektor-{sektor.nazwa}.zip"'
    return resp


# ---- Statyczne strony ----

def faq(request):
    return render(request, 'groby/faq.html')


def pomoc(request):
    return render(request, 'groby/pomoc.html')


def media_kit(request):
    return render(request, 'groby/media_kit.html')


# ---- Listy do zmarłych ----

@require_POST
def dodaj_list(request, osoba_id):
    if _antybot(request):
        return redirect('groby:osoba_detail', pk=osoba_id)
    osoba = get_object_or_404(Osoba, pk=osoba_id)
    tresc = (request.POST.get('tresc') or '').strip()
    if not tresc:
        return redirect('groby:osoba_detail', pk=osoba_id)
    List.objects.create(
        osoba=osoba,
        autor_user=request.user if request.user.is_authenticated else None,
        autor_imie=(request.POST.get('autor_imie') or '').strip()[:100],
        tresc=tresc[:5000],
        publiczny=request.POST.get('publiczny') == 'on',
    )
    messages.success(request, '✉️ List wysłany. Po akceptacji może pojawić się publicznie.')
    return redirect('groby:osoba_detail', pk=osoba_id)


def sciana_listow(request):
    listy = List.objects.filter(zaakceptowany=True, publiczny=True).select_related('osoba__grob__sektor', 'autor_user').order_by('-data_dodania')
    page = Paginator(listy, 20).get_page(request.GET.get('page'))
    return render(request, 'groby/sciana_listow.html', {'page': page})


# ---- Quiz ----

def quiz(request):
    import random
    pytania = list(PytanieQuiz.objects.filter(aktywne=True))
    random.shuffle(pytania)
    pytania = pytania[:10]
    return render(request, 'groby/quiz.html', {'pytania': pytania})


# ---- Forum ----

def forum_grobu(request, grob_id):
    grob = get_object_or_404(Grob.objects.select_related('sektor'), pk=grob_id)
    watki = grob.watki.select_related('autor').prefetch_related('posty')
    return render(request, 'groby/forum.html', {'grob': grob, 'watki': watki})


@login_required
@require_POST
def utworz_watek(request, grob_id):
    grob = get_object_or_404(Grob, pk=grob_id)
    tytul = (request.POST.get('tytul') or '').strip()[:200]
    tresc = (request.POST.get('tresc') or '').strip()
    if not tytul or not tresc:
        return redirect('groby:forum_grobu', grob_id=grob_id)
    watek = WatekForum.objects.create(grob=grob, autor=request.user, tytul=tytul)
    PostForum.objects.create(watek=watek, autor=request.user, tresc=tresc[:5000], zaakceptowany=True)
    return redirect('groby:forum_grobu', grob_id=grob_id)


@login_required
@require_POST
def dodaj_post(request, watek_id):
    watek = get_object_or_404(WatekForum, pk=watek_id)
    tresc = (request.POST.get('tresc') or '').strip()
    if tresc:
        PostForum.objects.create(watek=watek, autor=request.user, tresc=tresc[:5000])
        from django.utils import timezone
        watek.data_ostatniego_postu = timezone.now()
        watek.save(update_fields=['data_ostatniego_postu'])
    return redirect('groby:forum_grobu', grob_id=watek.grob_id)


# ---- Audit rollback ----

@login_required
def cofnij_zmiane(request, pk):
    if not request.user.is_staff:
        return redirect('admin:login')
    h = get_object_or_404(HistoriaZmian, pk=pk)
    if h.akcja != 'zmieniono' or not h.pola:
        messages.error(request, 'Można cofnąć tylko zmiany pól.')
        return redirect('groby:historia')
    Klasa = Grob if h.model == 'Grob' else (Osoba if h.model == 'Osoba' else None)
    if not Klasa:
        return redirect('groby:historia')
    obj = Klasa.objects.filter(pk=h.obiekt_id).first()
    if not obj:
        messages.error(request, f'{h.model} #{h.obiekt_id} już nie istnieje.')
        return redirect('groby:historia')
    if request.method == 'POST':
        for pole, dane in h.pola.items():
            try:
                setattr(obj, pole, dane.get('przed'))
            except Exception:
                pass
        try:
            obj.save()
            messages.success(request, f'Cofnięto zmiany w {h.model}#{h.obiekt_id}.')
        except Exception as e:
            messages.error(request, f'Błąd: {e}')
        return redirect('groby:historia')
    return render(request, 'groby/cofnij_zmiane.html', {'h': h, 'obj': obj})


# ---- Donation page ----

def donate(request):
    return render(request, 'groby/donate.html')


# ---- Apple Wallet pass (skeleton) ----

def wallet_pass(request, pk):
    """Generuje JSON dla Apple Wallet (uproszczony, bez podpisywania)."""
    grob = get_object_or_404(Grob.objects.select_related('sektor').prefetch_related('osoby'), pk=pk)
    pass_data = {
        'formatVersion': 1,
        'passTypeIdentifier': 'pass.cmentarz.szydlow',
        'serialNumber': str(grob.pk),
        'teamIdentifier': 'PLACEHOLDER',
        'organizationName': 'Cmentarz Szydłów',
        'description': f'Sektor {grob.sektor.nazwa} · grób {grob.numer}',
        'generic': {
            'primaryFields': [{'key': 'lokalizacja', 'label': 'Lokalizacja', 'value': f'Sektor {grob.sektor.nazwa}/{grob.numer}'}],
            'secondaryFields': [{'key': 'osoby', 'label': 'Pochowani', 'value': ', '.join(f'{o.imie} {o.nazwisko}' for o in grob.osoby.all()[:3])}],
            'backFields': [{'key': 'url', 'label': 'Strona', 'value': request.build_absolute_uri(reverse('groby:grob_detail', args=[grob.pk]))}],
        },
        'barcodes': [{'message': request.build_absolute_uri(reverse('groby:grob_detail', args=[grob.pk])),
                      'format': 'PKBarcodeFormatQR', 'messageEncoding': 'iso-8859-1'}],
    }
    return JsonResponse(pass_data)


# ---- Onboarding ----

@login_required
@require_POST
def zakoncz_onboarding(request):
    profil_obj, _ = Profil.objects.get_or_create(user=request.user)
    profil_obj.onboarding_zakonczony = True
    profil_obj.save(update_fields=['onboarding_zakonczony'])
    return JsonResponse({'ok': True})


def manifest(request):
    data = {
        "name": "Informator Cmentarny — Szydłów",
        "short_name": "Szydłów",
        "description": "Cmentarz parafialny w Szydłowie — wyszukiwarka grobów.",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#fbf9f4",
        "theme_color": "#2e4430",
        "lang": "pl",
        "icons": [
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png"},
        ],
    }
    return JsonResponse(data)


def service_worker(request):
    sw = """
const CACHE = 'groby-v1';
const ASSETY = ['/', '/szukaj/', '/sektory/', '/o-cmentarzu/', '/manifest.webmanifest'];
self.addEventListener('install', e => {
    e.waitUntil(caches.open(CACHE).then(c => c.addAll(ASSETY)).then(() => self.skipWaiting()));
});
self.addEventListener('activate', e => {
    e.waitUntil(caches.keys().then(ks => Promise.all(ks.filter(k => k !== CACHE).map(k => caches.delete(k)))).then(() => self.clients.claim()));
});
self.addEventListener('fetch', e => {
    if (e.request.method !== 'GET') return;
    const u = new URL(e.request.url);
    if (u.pathname.startsWith('/admin') || u.pathname.startsWith('/staff') || u.pathname.startsWith('/api')) return;
    e.respondWith(
        caches.match(e.request).then(c => c || fetch(e.request).then(r => {
            if (r.ok && (r.type === 'basic' || r.type === 'cors')) {
                const kopia = r.clone();
                caches.open(CACHE).then(cc => cc.put(e.request, kopia));
            }
            return r;
        }).catch(() => caches.match('/')))
    );
});
"""
    return HttpResponse(sw, content_type='application/javascript')


def grob_qr(request, pk):
    import qrcode
    grob = get_object_or_404(Grob, pk=pk)
    url = request.build_absolute_uri(reverse('groby:grob_detail', args=[grob.pk]))
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return HttpResponse(buf.getvalue(), content_type='image/png')


def qr_naklejki(request):
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    sektor_id = request.GET.get('sektor', '').strip()
    qs = Grob.objects.select_related('sektor').order_by('sektor__nazwa', 'rzad', 'numer')
    if sektor_id:
        qs = qs.filter(sektor_id=sektor_id)
    return render(request, 'groby/qr_naklejki.html', {
        'groby': qs,
        'sektory': Sektor.objects.order_by('nazwa'),
        'wybrany_sektor': sektor_id,
    })


def zglos_poprawke(request, cel, pk):
    if request.method != 'POST':
        return redirect('groby:home')
    if _antybot(request):
        messages.error(request, 'Wykryto podejrzaną aktywność. Spróbuj ponownie.')
        return redirect(request.META.get('HTTP_REFERER', '/'))
    tresc = (request.POST.get('tresc') or '').strip()
    if not tresc:
        messages.error(request, 'Treść zgłoszenia nie może być pusta.')
        return redirect(request.META.get('HTTP_REFERER', '/'))
    autor_imie = (request.POST.get('autor_imie') or '').strip()[:100]
    autor_email = (request.POST.get('autor_email') or '').strip()[:200]
    z = Zgloszenie(
        tresc=tresc,
        autor_imie=autor_imie,
        autor_email=autor_email,
        autor_user=request.user if request.user.is_authenticated else None,
    )
    if cel == 'grob':
        z.grob = get_object_or_404(Grob, pk=pk)
    elif cel == 'osoba':
        z.osoba = get_object_or_404(Osoba, pk=pk)
    else:
        return redirect('groby:home')
    z.save()
    messages.success(request, 'Dziękujemy! Zgłoszenie zostało zarejestrowane.')
    return redirect(request.META.get('HTTP_REFERER', '/'))


def eksport_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from pathlib import Path

    osoby = Osoba.objects.select_related('grob', 'grob__sektor')
    query = request.GET.get('q', '').strip()
    sektor_id = request.GET.get('sektor', '').strip()
    typ = request.GET.get('typ', '').strip()
    rok_od = request.GET.get('rok_od', '').strip()
    rok_do = request.GET.get('rok_do', '').strip()
    if query:
        osoby = osoby.filter(Q(nazwisko__icontains=query) | Q(imie__icontains=query) | Q(nazwisko_rodowe__icontains=query))
    if sektor_id:
        osoby = osoby.filter(grob__sektor_id=sektor_id)
    if typ:
        osoby = osoby.filter(grob__typ=typ)
    if rok_od.isdigit():
        osoby = osoby.filter(data_smierci__year__gte=int(rok_od))
    if rok_do.isdigit():
        osoby = osoby.filter(data_smierci__year__lte=int(rok_do))
    osoby = osoby.order_by('nazwisko', 'imie')[:2000]

    czcionka_nazwa = 'Helvetica'
    for kandydat in (r'C:\Windows\Fonts\arial.ttf', r'C:\Windows\Fonts\calibri.ttf'):
        if Path(kandydat).exists():
            try:
                pdfmetrics.registerFont(TTFont('Polska', kandydat))
                czcionka_nazwa = 'Polska'
                break
            except Exception:
                pass

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    style_h1 = ParagraphStyle('h1', fontName=czcionka_nazwa, fontSize=18, leading=22, spaceAfter=6)
    style_meta = ParagraphStyle('meta', fontName=czcionka_nazwa, fontSize=9, textColor=colors.grey, spaceAfter=12)
    style_n = ParagraphStyle('n', fontName=czcionka_nazwa, fontSize=9, leading=11)
    elementy = []
    elementy.append(Paragraph('Księga zmarłych — cmentarz w Szydłowie', style_h1))
    filtry = []
    if query: filtry.append(f'fraza: „{query}"')
    if sektor_id: filtry.append(f'sektor: {Sektor.objects.filter(pk=sektor_id).values_list("nazwa", flat=True).first() or "?"}')
    if typ: filtry.append(f'typ: {dict(Grob.TYP_CHOICES).get(typ, typ)}')
    if rok_od: filtry.append(f'od {rok_od}')
    if rok_do: filtry.append(f'do {rok_do}')
    elementy.append(Paragraph('Filtry: ' + ('; '.join(filtry) if filtry else 'brak') + f'.   Wyników: {len(osoby)}.', style_meta))

    naglowek = ['Lp.', 'Nazwisko i imię', 'Daty', 'Lokalizacja']
    dane = [naglowek]
    for i, o in enumerate(osoby, start=1):
        nazwa = f'{o.nazwisko} {o.imie}'
        if o.nazwisko_rodowe:
            nazwa += f' (z d. {o.nazwisko_rodowe})'
        ur = o.data_urodzenia.strftime('%Y-%m-%d') if o.data_urodzenia else '?'
        sm = o.data_smierci.strftime('%Y-%m-%d') if o.data_smierci else '?'
        lok = f'sekt. {o.grob.sektor.nazwa}'
        if o.grob.rzad:
            lok += f', rz. {o.grob.rzad}'
        lok += f', nr {o.grob.numer}'
        dane.append([
            Paragraph(str(i), style_n),
            Paragraph(nazwa, style_n),
            Paragraph(f'{ur} — {sm}', style_n),
            Paragraph(lok, style_n),
        ])
    tab = Table(dane, colWidths=[1.0 * cm, 7.5 * cm, 4.0 * cm, 5.0 * cm], repeatRows=1)
    tab.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), czcionka_nazwa),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2e4430')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f6ef')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    elementy.append(tab)
    doc.build(elementy)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="ksiega-zmarlych.pdf"'
    return resp


def rejestracja(request):
    if request.user.is_authenticated:
        return redirect('groby:profil')
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            Profil.objects.get_or_create(user=user)
            login(request, user)
            messages.success(request, 'Konto utworzone.')
            return redirect('groby:profil')
    else:
        form = UserCreationForm()
    return render(request, 'groby/rejestracja.html', {'form': form})


@login_required
def profil(request):
    profil_obj, _ = Profil.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        profil_obj.pokrewienstwo = (request.POST.get('pokrewienstwo') or '').strip()[:200]
        profil_obj.save(update_fields=['pokrewienstwo'])
        messages.success(request, 'Profil zapisany.')
        return redirect('groby:profil')
    return render(request, 'groby/profil.html', {
        'profil': profil_obj,
        'obserwowane_groby': profil_obj.obserwowane_groby.select_related('sektor').prefetch_related('osoby'),
        'obserwowane_osoby': profil_obj.obserwowane_osoby.select_related('grob__sektor'),
        'moje_zgloszenia': Zgloszenie.objects.filter(autor_user=request.user).select_related('grob__sektor', 'osoba__grob__sektor')[:20],
        'zapisane_szukania': request.user.zapisane_szukania.all()[:20],
        'moje_odznaki': request.user.odznaki.select_related('odznaka'),
    })


@login_required
@require_POST
def przelacz_obserwacje(request, cel, pk):
    profil_obj, _ = Profil.objects.get_or_create(user=request.user)
    if cel == 'grob':
        obj = get_object_or_404(Grob, pk=pk)
        if profil_obj.obserwowane_groby.filter(pk=obj.pk).exists():
            profil_obj.obserwowane_groby.remove(obj)
        else:
            profil_obj.obserwowane_groby.add(obj)
    elif cel == 'osoba':
        obj = get_object_or_404(Osoba, pk=pk)
        if profil_obj.obserwowane_osoby.filter(pk=obj.pk).exists():
            profil_obj.obserwowane_osoby.remove(obj)
        else:
            profil_obj.obserwowane_osoby.add(obj)
    return redirect(request.META.get('HTTP_REFERER', '/'))


def dashboard_staff(request):
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    metryki = {
        'zgloszenia_nowe': Zgloszenie.objects.filter(status='nowe').count(),
        'wspomnienia_oczekujace': Wspomnienie.objects.filter(status='oczekuje').count(),
        'groby_bez_pozycji': Grob.objects.filter(plan_x__isnull=True).count(),
        'groby_bez_osob': Grob.objects.annotate(_n=Count('osoby')).filter(_n=0).count(),
        'osoby_bez_dat': Osoba.objects.filter(data_smierci__isnull=True).count(),
        'liczba_zdjec': Zdjecie.objects.count(),
        'aktywne_swiece_24h': Swieca.objects.filter(data_zapalenia__gte=_now_minus_hours(24)).count(),
        'liczba_userow': __import__('django.contrib.auth', fromlist=['get_user_model']).get_user_model().objects.count(),
        'ostatnie_zmiany': HistoriaZmian.objects.select_related('user').order_by('-data')[:10],
        'najnowsze_zgloszenia': Zgloszenie.objects.select_related('grob__sektor', 'osoba').order_by('-data_dodania')[:5],
        'najnowsze_wspomnienia': Wspomnienie.objects.filter(status='oczekuje').select_related('osoba')[:5],
    }
    return render(request, 'groby/dashboard.html', metryki)


def _now_minus_hours(h):
    from datetime import timedelta
    from django.utils import timezone
    return timezone.now() - timedelta(hours=h)


def duplikaty(request):
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    klucze = defaultdict(list)
    for o in Osoba.objects.select_related('grob__sektor').all():
        klucz = (o.imie.lower().strip(), o.nazwisko.lower().strip(),
                 o.data_smierci.year if o.data_smierci else None)
        klucze[klucz].append(o)
    grupy = []
    for klucz, lista in klucze.items():
        if len(lista) > 1:
            grupy.append({
                'imie': klucz[0].title(),
                'nazwisko': klucz[1].title(),
                'rok': klucz[2],
                'osoby': sorted(lista, key=lambda o: o.pk),
            })
    grupy.sort(key=lambda g: (g['nazwisko'], g['imie']))
    return render(request, 'groby/duplikaty.html', {
        'grupy': grupy,
    })


def historia_zmian(request):
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    qs = HistoriaZmian.objects.select_related('user').order_by('-data')
    model = request.GET.get('model', '').strip()
    if model:
        qs = qs.filter(model=model)
    page = Paginator(qs, 50).get_page(request.GET.get('page'))
    return render(request, 'groby/historia.html', {
        'page': page,
        'model_filtr': model,
    })


def search_analytics(request):
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    from django.db.models.functions import TruncDate
    top_frazy = (WyszukiwanieLog.objects.values('fraza')
                 .annotate(n=Count('id'), wyniki=Count('id'))
                 .order_by('-n')[:50])
    bez_wynikow = (WyszukiwanieLog.objects.filter(liczba_wynikow=0)
                   .values('fraza').annotate(n=Count('id')).order_by('-n')[:30])
    dziennie = (WyszukiwanieLog.objects.annotate(d=TruncDate('data'))
                .values('d').annotate(n=Count('id')).order_by('-d')[:30])
    return render(request, 'groby/search_analytics.html', {
        'top_frazy': list(top_frazy),
        'bez_wynikow': list(bez_wynikow),
        'dziennie': list(dziennie),
    })


def dronowe(request):
    zdjecia = ZdjecieDronowe.objects.select_related('sektor').order_by('kolejnosc', '-data_wykonania')
    return render(request, 'groby/dronowe.html', {'zdjecia': zdjecia})


def cmentarz_w_czasie(request):
    """Liczba pochówków per dekada (na podstawie data_smierci)."""
    from collections import Counter
    licznik = Counter()
    for o in Osoba.objects.exclude(data_smierci__isnull=True).only('data_smierci'):
        d = o.data_smierci.year // 10 * 10
        licznik[d] += 1
    dekady = sorted(licznik.items())
    return render(request, 'groby/cmentarz_w_czasie.html', {
        'dekady': dekady,
        'dekady_json': json.dumps(dekady),
    })


def powracajacy(request):
    """Lista grobów najczęściej edytowanych w ciągu ostatnich 90 dni."""
    from django.utils import timezone
    from datetime import timedelta
    od = timezone.now() - timedelta(days=90)
    naj = (HistoriaZmian.objects.filter(model='Grob', data__gte=od)
           .values('obiekt_id', 'obiekt_repr')
           .annotate(n=Count('id')).order_by('-n')[:30])
    return render(request, 'groby/powracajacy.html', {'naj': naj})


def konkurs_lista(request):
    aktywne = KonkursFoto.objects.filter(aktywny=True).order_by('-data_start')
    archiwalne = KonkursFoto.objects.filter(aktywny=False).order_by('-data_start')[:10]
    return render(request, 'groby/konkurs_lista.html', {
        'aktywne': aktywne,
        'archiwalne': archiwalne,
    })


def konkurs_detail(request, pk):
    konkurs = get_object_or_404(KonkursFoto, pk=pk)
    zgloszenia = (konkurs.zgloszenia_foto.filter(zaakceptowane=True)
                  .annotate(liczba_glosow=Count('glosy'))
                  .order_by('-liczba_glosow', '-data_dodania'))
    return render(request, 'groby/konkurs_detail.html', {
        'konkurs': konkurs,
        'zgloszenia': zgloszenia,
    })


def konkurs_zgloszenie(request, pk):
    konkurs = get_object_or_404(KonkursFoto, pk=pk, aktywny=True)
    if request.method == 'POST':
        if _antybot(request):
            return redirect('groby:konkurs_lista')
        plik = request.FILES.get('plik')
        tytul = request.POST.get('tytul', '').strip()[:200]
        autor_imie = request.POST.get('autor_imie', '').strip()[:100]
        if plik:
            ZgloszenieKonkursowe.objects.create(
                konkurs=konkurs,
                autor=request.user if request.user.is_authenticated else None,
                autor_imie=autor_imie,
                plik=plik,
                tytul=tytul,
            )
            messages.success(request, 'Zgłoszenie wysłane — czeka na akceptację.')
            return redirect('groby:konkurs_detail', pk=konkurs.pk)
    return render(request, 'groby/konkurs_zgloszenie.html', {'konkurs': konkurs})


@require_POST
def konkurs_glosuj(request, pk):
    z = get_object_or_404(ZgloszenieKonkursowe, pk=pk, zaakceptowane=True)
    ip = _hash_ip(request)
    obj, utworzono = GlosKonkursowy.objects.get_or_create(
        zgloszenie=z, ip_hash=ip,
        defaults={'user': request.user if request.user.is_authenticated else None},
    )
    if utworzono:
        messages.success(request, 'Głos zapisany.')
    else:
        messages.info(request, 'Już oddałeś głos na to zdjęcie.')
    return redirect('groby:konkurs_detail', pk=z.konkurs_id)


def cmentarna_ksiega_pdf(request):
    """Pełna księga: wszystkie groby z osobami, posortowane sektor/rząd/numer."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title='Księga Cmentarna')
    style = getSampleStyleSheet()
    flow = [Paragraph('Księga Cmentarna — Szydłów', style['Title']), Spacer(1, 12)]
    for sektor in Sektor.objects.order_by('nazwa'):
        flow.append(Paragraph(f'Sektor {sektor.nazwa}', style['Heading2']))
        groby = (Grob.objects.filter(sektor=sektor)
                 .prefetch_related('osoby').order_by('rzad', 'numer'))
        for g in groby:
            naglowek = f'Grób {g.rzad}/{g.numer}'
            flow.append(Paragraph(naglowek, style['Heading4']))
            for o in g.osoby.all():
                txt = f'{o.imie or "?"} {o.nazwisko or ""}'.strip()
                if o.data_urodzenia:
                    txt += f', ur. {o.data_urodzenia}'
                if o.data_smierci:
                    txt += f', zm. {o.data_smierci}'
                flow.append(Paragraph(txt, style['Normal']))
            flow.append(Spacer(1, 4))
        flow.append(PageBreak())
    doc.build(flow)
    pdf = buf.getvalue()
    buf.close()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="ksiega_cmentarna.pdf"'
    return resp


def family_book_pdf(request, osoba_id):
    osoba = get_object_or_404(Osoba, pk=osoba_id)
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f'Księga rodu {osoba.nazwisko}')
    style = getSampleStyleSheet()
    flow = [Paragraph(f'Księga rodu {osoba.nazwisko or osoba.imie}', style['Title']), Spacer(1, 12)]
    przodkowie = _zbierz_przodkow(osoba) if '_zbierz_przodkow' in globals() else []
    for o in przodkowie or [osoba]:
        flow.append(Paragraph(f'<b>{o.imie or ""} {o.nazwisko or ""}</b>', style['Heading3']))
        meta = []
        if o.data_urodzenia: meta.append(f'ur. {o.data_urodzenia}')
        if o.data_smierci: meta.append(f'zm. {o.data_smierci}')
        if meta:
            flow.append(Paragraph(', '.join(meta), style['Normal']))
        if o.biogram:
            flow.append(Paragraph(o.biogram, style['Normal']))
        flow.append(Spacer(1, 8))
    doc.build(flow)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="ksiega_rodu_{osoba.nazwisko or "osoba"}.pdf"'
    return resp


def folder_turystyczny_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title='Folder turystyczny')
    style = getSampleStyleSheet()
    flow = [
        Paragraph('Cmentarz Parafialny w Szydłowie', style['Title']),
        Spacer(1, 8),
        Paragraph('Przewodnik turystyczny', style['Heading2']),
        Spacer(1, 12),
        Paragraph(
            'Szydłów to średniowieczne miasteczko z zachowanymi murami obronnymi z XIV wieku. '
            'Cmentarz parafialny stanowi cenny zabytek lokalnej historii — spoczywają tu '
            'pokolenia mieszkańców, w tym postacie zasłużone dla regionu.',
            style['Normal']),
        Spacer(1, 8),
    ]
    flow.append(Paragraph('Najważniejsze trasy', style['Heading3']))
    for trasa in Trasa.objects.order_by('-opublikowana', 'nazwa')[:5]:
        flow.append(Paragraph(f'<b>{trasa.nazwa}</b>', style['Normal']))
        if trasa.opis:
            flow.append(Paragraph(trasa.opis, style['Normal']))
        flow.append(Spacer(1, 4))
    doc.build(flow)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="folder_turystyczny.pdf"'
    return resp


def ulotka_pdf(request):
    """Ulotka edukacyjna: zasady na cmentarzu, kontakt, QR do strony."""
    from reportlab.lib.pagesizes import A5
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A5, title='Ulotka — Cmentarz w Szydłowie')
    style = getSampleStyleSheet()
    flow = [
        Paragraph('Cmentarz Parafialny w Szydłowie', style['Title']),
        Spacer(1, 6),
        Paragraph('Zasady zachowania', style['Heading3']),
        Paragraph('— Zachowaj ciszę i powagę miejsca.', style['Normal']),
        Paragraph('— Nie pal i nie spożywaj alkoholu.', style['Normal']),
        Paragraph('— Sprzątaj po sobie i swoich bliskich.', style['Normal']),
        Paragraph('— Świece zapalaj z rozwagą.', style['Normal']),
        Spacer(1, 8),
        Paragraph('Strona internetowa', style['Heading3']),
        Paragraph('Pełna baza grobów, biogramów i zdjęć dostępna online.', style['Normal']),
        Paragraph('Zeskanuj QR przy bramie głównej.', style['Normal']),
    ]
    doc.build(flow)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="ulotka.pdf"'
    return resp


def prometheus_metrics(request):
    """Minimalny endpoint /metrics zgodny z Prometheus."""
    metrics = [
        f'# HELP groby_total Liczba grobów',
        f'# TYPE groby_total gauge',
        f'groby_total {Grob.objects.count()}',
        f'# HELP osoby_total Liczba osób',
        f'# TYPE osoby_total gauge',
        f'osoby_total {Osoba.objects.count()}',
        f'# HELP zdjecia_total Liczba zdjęć',
        f'# TYPE zdjecia_total gauge',
        f'zdjecia_total {Zdjecie.objects.count()}',
        f'# HELP wspomnienia_total Liczba wspomnień',
        f'# TYPE wspomnienia_total gauge',
        f'wspomnienia_total {Wspomnienie.objects.count()}',
        f'# HELP swiece_total Liczba zapalonych świec',
        f'# TYPE swiece_total counter',
        f'swiece_total {Swieca.objects.count()}',
        f'# HELP wyszukiwania_total Liczba wyszukiwań',
        f'# TYPE wyszukiwania_total counter',
        f'wyszukiwania_total {WyszukiwanieLog.objects.count()}',
    ]
    return HttpResponse('\n'.join(metrics) + '\n', content_type='text/plain; version=0.0.4')


def live_stats_json(request):
    """JSON dla widgetu live — odświeżany co 30 s przez fetch."""
    from django.utils import timezone
    from datetime import timedelta
    od = timezone.now() - timedelta(hours=24)
    return JsonResponse({
        'osoby': Osoba.objects.count(),
        'groby': Grob.objects.count(),
        'swiece_24h': Swieca.objects.filter(data_zapalenia__gte=od).count(),
        'wspomnienia_24h': Wspomnienie.objects.filter(data_dodania__gte=od).count(),
        'wyszukiwania_24h': WyszukiwanieLog.objects.filter(data__gte=od).count(),
    })


def dzis_rocznica_json(request):
    """JSON: osoby z rocznicą śmierci/urodzin dziś."""
    from django.utils import timezone
    today = timezone.localdate()
    rocznice = Osoba.objects.filter(
        data_smierci__month=today.month, data_smierci__day=today.day
    ).exclude(data_smierci=today).order_by('nazwisko')[:20]
    urodziny = Osoba.objects.filter(
        data_urodzenia__month=today.month, data_urodzenia__day=today.day
    ).exclude(data_urodzenia=today).order_by('nazwisko')[:20]
    return JsonResponse({
        'rocznice_smierci': [
            {'id': o.pk, 'imie_nazwisko': f'{o.imie or ""} {o.nazwisko or ""}'.strip(),
             'lat': today.year - o.data_smierci.year}
            for o in rocznice
        ],
        'urodziny': [
            {'id': o.pk, 'imie_nazwisko': f'{o.imie or ""} {o.nazwisko or ""}'.strip(),
             'lat': today.year - o.data_urodzenia.year}
            for o in urodziny
        ],
    })


def ai_biogram(request, osoba_id):
    """Skeleton — wymaga klucza API i ręcznej weryfikacji."""
    osoba = get_object_or_404(Osoba, pk=osoba_id)
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    propozycja = None
    if request.method == 'POST':
        import os
        klucz = os.environ.get('ANTHROPIC_API_KEY') or os.environ.get('OPENAI_API_KEY')
        if not klucz:
            messages.error(request, 'Brak klucza API (ANTHROPIC_API_KEY / OPENAI_API_KEY).')
        else:
            propozycja = (
                f'{osoba.imie or ""} {osoba.nazwisko or ""} '
                f'({osoba.data_urodzenia or "?"} – {osoba.data_smierci or "?"}). '
                f'[Wygenerowany szkielet biogramu — wymaga ręcznej weryfikacji historycznej.]'
            )
    return render(request, 'groby/ai_biogram.html', {
        'osoba': osoba,
        'propozycja': propozycja,
    })


# ----- Batch 90 -----


def diff_zmiany(request, pk):
    """Side-by-side diff dla wpisu HistoriaZmian."""
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    h = get_object_or_404(HistoriaZmian, pk=pk)
    return render(request, 'groby/diff_zmiany.html', {'h': h})


def time_lapse_grobu(request, pk):
    grob = get_object_or_404(Grob, pk=pk)
    zdjecia = grob.zdjecia.order_by('kolejnosc', 'pk')
    return render(request, 'groby/time_lapse.html', {'grob': grob, 'zdjecia': zdjecia})


HISTORIA_PL = [
    (1772, 1795, 'Rozbiory Polski'),
    (1830, 1831, 'Powstanie listopadowe'),
    (1846, 1846, 'Powstanie krakowskie / rzeź galicyjska'),
    (1863, 1864, 'Powstanie styczniowe'),
    (1914, 1918, 'I wojna światowa'),
    (1918, 1918, 'Odzyskanie niepodległości'),
    (1920, 1920, 'Bitwa Warszawska'),
    (1939, 1945, 'II wojna światowa'),
    (1945, 1989, 'PRL'),
    (1980, 1981, 'Solidarność'),
    (1989, 1989, 'Zmiana ustroju'),
    (2004, 2004, 'Wstąpienie do UE'),
]


def historic_events_overlay(request, osoba_id):
    """JSON: wydarzenia historyczne, które miały miejsce za życia osoby."""
    osoba = get_object_or_404(Osoba, pk=osoba_id)
    if not osoba.data_urodzenia:
        return JsonResponse({'wydarzenia': []})
    rok_ur = osoba.data_urodzenia.year
    rok_zm = osoba.data_smierci.year if osoba.data_smierci else 2026
    pasujace = [
        {'od': od, 'do': do, 'tytul': t}
        for od, do, t in HISTORIA_PL
        if not (do < rok_ur or od > rok_zm)
    ]
    return JsonResponse({
        'urodzenie': rok_ur,
        'smierc': rok_zm,
        'wydarzenia': pasujace,
    })


def _sesja_id(request):
    if not request.session.session_key:
        request.session.create()
    return request.session.session_key


def plan_zwiedzania(request):
    user = request.user if request.user.is_authenticated else None
    if user:
        plan = PlanZwiedzania.objects.filter(user=user).select_related('grob__sektor')
    else:
        plan = PlanZwiedzania.objects.filter(sesja_id=_sesja_id(request), user__isnull=True).select_related('grob__sektor')
    return render(request, 'groby/plan_zwiedzania.html', {'plan': plan})


@require_POST
def plan_dodaj(request, grob_id):
    grob = get_object_or_404(Grob, pk=grob_id)
    user = request.user if request.user.is_authenticated else None
    sesja = '' if user else _sesja_id(request)
    PlanZwiedzania.objects.get_or_create(user=user, sesja_id=sesja, grob=grob)
    messages.success(request, 'Dodano do planu zwiedzania.')
    return redirect(request.META.get('HTTP_REFERER') or 'groby:plan_zwiedzania')


@require_POST
def plan_oznacz(request, pk):
    p = get_object_or_404(PlanZwiedzania, pk=pk)
    user = request.user if request.user.is_authenticated else None
    if (user and p.user_id != user.pk) or (not user and p.sesja_id != request.session.session_key):
        return JsonResponse({'ok': False}, status=403)
    p.odwiedzony = not p.odwiedzony
    p.save(update_fields=['odwiedzony'])
    return redirect('groby:plan_zwiedzania')


@require_POST
def plan_usun(request, pk):
    p = get_object_or_404(PlanZwiedzania, pk=pk)
    user = request.user if request.user.is_authenticated else None
    if (user and p.user_id != user.pk) or (not user and p.sesja_id != request.session.session_key):
        return JsonResponse({'ok': False}, status=403)
    p.delete()
    return redirect('groby:plan_zwiedzania')


def heatmapa_swiec_json(request):
    """Punkty (x,y,liczba) dla heatmapy zapaleń świec."""
    from django.utils import timezone
    from datetime import timedelta
    od = timezone.now() - timedelta(days=30)
    qs = (Swieca.objects.filter(data_zapalenia__gte=od, osoba__grob__plan_x__isnull=False)
          .values('osoba__grob__plan_x', 'osoba__grob__plan_y')
          .annotate(n=Count('id')))
    pkty = [
        {'x': p['osoba__grob__plan_x'], 'y': p['osoba__grob__plan_y'], 'n': p['n']}
        for p in qs if p['osoba__grob__plan_x'] is not None
    ]
    return JsonResponse({'punkty': pkty})


def featured_tygodnia():
    """Zwraca aktualne wyróżnienie (osoba/grób/wpis) lub deterministycznie auto-wybiera."""
    from django.utils import timezone
    today = timezone.localdate()
    f = FeaturedTygodnia.objects.filter(aktywne=True, od__lte=today, do__gte=today).first()
    if f:
        return f
    osoby = Osoba.objects.exclude(biogram='').order_by('pk')
    if osoby.exists():
        idx = today.isocalendar()[1] % osoby.count()
        wybrana = osoby[idx]
        return {
            'kategoria': 'osoba',
            'tytul': f'{wybrana.imie or ""} {wybrana.nazwisko or ""}'.strip(),
            'opis': (wybrana.biogram or '')[:200],
            'osoba': wybrana,
        }
    return None


def importer_exif_info(request):
    """Strona pomocy: jak importować zdjęcia z EXIF GPS (link do management command)."""
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    return render(request, 'groby/importer_exif.html', {})


def stripe_donate(request):
    """Skeleton: tworzy Checkout session (wymaga STRIPE_SECRET_KEY)."""
    import os
    klucz = os.environ.get('STRIPE_SECRET_KEY', '')
    kwota = int(request.POST.get('kwota', '50') or '50')
    if not klucz:
        messages.error(request, 'Brak skonfigurowanego STRIPE_SECRET_KEY.')
        return redirect('groby:donate')
    return JsonResponse({
        'info': 'Stripe skeleton — w produkcji użyj stripe.checkout.Session.create() z success_url.',
        'kwota_pln': kwota,
        'klucz_publiczny_env': 'STRIPE_PUBLISHABLE_KEY',
    })


def audit_log_pdf(request):
    """Eksport HistoriaZmian do PDF (compliance/RODO)."""
    if not (request.user.is_authenticated and request.user.is_staff):
        return redirect('admin:login')
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title='Audit log')
    style = getSampleStyleSheet()
    flow = [Paragraph('Dziennik zmian — eksport zgodności (RODO)', style['Title']), Spacer(1, 8)]
    qs = HistoriaZmian.objects.select_related('user').order_by('-data')[:1000]
    dane = [['Data', 'Akcja', 'Model', 'Obiekt', 'Użytkownik']]
    for h in qs:
        dane.append([
            h.data.strftime('%Y-%m-%d %H:%M'),
            h.akcja,
            h.model,
            (h.obiekt_repr or '')[:60],
            (h.user.username if h.user else '—')[:30],
        ])
    t = Table(dane, repeatRows=1)
    t.setStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5b6e51')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ])
    flow.append(t)
    doc.build(flow)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="audit_log.pdf"'
    return resp


def lista_wpisow_z_tagami(request):
    """Rozszerzenie /postacie/ o filtr po tagach (przekierowanie z parametrem)."""
    tag = request.GET.get('tag', '').strip()
    qs = Wpis.objects.filter(opublikowany=True).order_by('-data_publikacji')
    if tag:
        qs = qs.filter(tagi_tresci__slug=tag)
    return render(request, 'groby/wpisy_lista.html', {
        'wpisy': qs[:50],
        'tagi_wpisow': TagWpisu.objects.all(),
        'aktywny_tag': tag,
    })


# ===== Batch 91 =====


@login_required
def powiadomienia_lista(request):
    qs = Powiadomienie.objects.filter(user=request.user)[:200]
    Powiadomienie.objects.filter(user=request.user, przeczytane=False).update(przeczytane=True)
    return render(request, 'groby/powiadomienia.html', {'powiadomienia': qs})


@login_required
def powiadomienia_count(request):
    n = Powiadomienie.objects.filter(user=request.user, przeczytane=False).count()
    return JsonResponse({'nieprzeczytane': n})


@require_POST
@login_required
def powiadomienie_oznacz(request, pk):
    p = get_object_or_404(Powiadomienie, pk=pk, user=request.user)
    p.przeczytane = True
    p.save(update_fields=['przeczytane'])
    return JsonResponse({'ok': True})


@login_required
def opieka_zgloszenie(request, grob_id):
    grob = get_object_or_404(Grob, pk=grob_id)
    if request.method == 'POST':
        if _antybot(request):
            return JsonResponse({'ok': False, 'powod': 'antybot'}, status=400)
        relacja = (request.POST.get('relacja') or '').strip()[:100]
        motywacja = (request.POST.get('motywacja') or '').strip()[:2000]
        op, created = OpiekunGrobu.objects.get_or_create(
            grob=grob, user=request.user,
            defaults={'relacja': relacja, 'motywacja': motywacja},
        )
        if created:
            messages.success(request, 'Zgłoszenie wysłane — staff je rozpatrzy.')
        else:
            messages.info(request, 'Już wysłałeś wcześniej zgłoszenie do tego grobu.')
        return redirect('groby:grob_detail', pk=grob.pk)
    return render(request, 'groby/opieka_form.html', {'grob': grob, 'antybot_html': _pole_antybot_html()})


def opiekunowie_lista(request):
    """Publiczna lista aktywnych opiekunów grobów (transparentność)."""
    qs = (OpiekunGrobu.objects.filter(status='aktywny')
          .select_related('grob__sektor', 'user')
          .order_by('grob__sektor__nazwa', 'grob__numer'))
    return render(request, 'groby/opiekunowie.html', {'opiekunowie': qs})


@login_required
@require_POST
def opieka_zrezygnuj(request, pk):
    op = get_object_or_404(OpiekunGrobu, pk=pk, user=request.user)
    op.status = 'zakonczony'
    op.save(update_fields=['status', 'data_zmiany'])
    messages.success(request, 'Zrezygnowano z opieki.')
    return redirect('groby:grob_detail', pk=op.grob_id)


@login_required
def prywatne_notatki(request, osoba_id):
    osoba = get_object_or_404(Osoba, pk=osoba_id)
    if request.method == 'POST':
        if _antybot(request):
            return JsonResponse({'ok': False}, status=400)
        tresc = (request.POST.get('tresc') or '').strip()
        if tresc:
            PrywatnaNotatka.objects.create(user=request.user, osoba=osoba, tresc=tresc[:5000])
            messages.success(request, 'Notatka zapisana.')
        return redirect('groby:prywatne_notatki', osoba_id=osoba.pk)
    notatki = PrywatnaNotatka.objects.filter(user=request.user, osoba=osoba)
    return render(request, 'groby/prywatne_notatki.html', {
        'osoba': osoba,
        'notatki': notatki,
        'antybot_html': _pole_antybot_html(),
    })


@login_required
@require_POST
def prywatna_notatka_usun(request, pk):
    n = get_object_or_404(PrywatnaNotatka, pk=pk, user=request.user)
    osoba_id = n.osoba_id
    n.delete()
    return redirect('groby:prywatne_notatki', osoba_id=osoba_id)


def niedawno_zmarli(request):
    from django.utils import timezone
    from datetime import timedelta
    try:
        dni = int(request.GET.get('dni', '90'))
    except ValueError:
        dni = 90
    dni = max(7, min(dni, 365))
    dzis = timezone.localdate()
    od = dzis - timedelta(days=dni)
    qs = (Osoba.objects.filter(data_smierci__gte=od, data_smierci__lte=dzis)
          .select_related('grob__sektor').order_by('-data_smierci'))
    return render(request, 'groby/niedawno_zmarli.html', {
        'osoby': qs,
        'dni': dni,
        'od': od,
        'do': dzis,
    })


def slownik_lista(request):
    q = (request.GET.get('q') or '').strip()
    kat = (request.GET.get('kat') or '').strip()
    qs = HasloSlownik.objects.all()
    if q:
        qs = qs.filter(Q(haslo__icontains=q) | Q(skrot__icontains=q) | Q(tresc__icontains=q))
    if kat:
        qs = qs.filter(kategoria=kat)
    return render(request, 'groby/slownik_lista.html', {
        'hasla': qs,
        'q': q,
        'kat': kat,
        'kategorie': HasloSlownik.KATEGORIA_CHOICES,
    })


def slownik_haslo(request, slug):
    haslo = get_object_or_404(HasloSlownik, slug=slug)
    powiazane = HasloSlownik.objects.filter(kategoria=haslo.kategoria).exclude(pk=haslo.pk)[:8]
    return render(request, 'groby/slownik_haslo.html', {
        'haslo': haslo,
        'zrodla': [s.strip() for s in (haslo.zrodla or '').splitlines() if s.strip()],
        'powiazane': powiazane,
    })


def statystyki_dlugowiecznosci(request):
    """Średnia długość życia per sektor i per dekada urodzenia."""
    osoby = Osoba.objects.exclude(data_urodzenia=None).exclude(data_smierci=None).select_related('grob__sektor')
    per_sektor = defaultdict(list)
    per_dekada = defaultdict(list)
    for o in osoby:
        wiek = o.wiek
        if wiek is None or wiek < 0 or wiek > 130:
            continue
        if o.grob and o.grob.sektor:
            per_sektor[o.grob.sektor.nazwa].append(wiek)
        dek = (o.data_urodzenia.year // 10) * 10
        per_dekada[dek].append(wiek)
    sektory = sorted(
        [(nazwa, round(sum(w) / len(w), 1), len(w), min(w), max(w)) for nazwa, w in per_sektor.items() if w],
        key=lambda x: -x[1],
    )
    dekady = sorted(
        [(d, round(sum(w) / len(w), 1), len(w)) for d, w in per_dekada.items() if w],
        key=lambda x: x[0],
    )
    wszyscy = [w for ws in per_sektor.values() for w in ws]
    srednia_calosc = round(sum(wszyscy) / len(wszyscy), 1) if wszyscy else 0
    return render(request, 'groby/statystyki_dlugowiecznosci.html', {
        'sektory': sektory,
        'dekady': dekady,
        'srednia_calosc': srednia_calosc,
        'razem': len(wszyscy),
    })


# ===== Batch 92 =====


def etykiety_lista(request):
    """Lista wszystkich etykiet osób z liczbą oznaczonych."""
    etykiety = EtykietaOsoby.objects.annotate(n=Count('osoby')).order_by('-n', 'nazwa')
    return render(request, 'groby/etykiety_lista.html', {'etykiety': etykiety})


def etykieta_detail(request, slug):
    et = get_object_or_404(EtykietaOsoby, slug=slug)
    osoby = et.osoby.select_related('grob__sektor').order_by('nazwisko', 'imie')
    return render(request, 'groby/etykieta_detail.html', {'et': et, 'osoby': osoby})


def statystyki_imion(request):
    """Top imiona m/k, top nazwiska, popularność imion w dekadach urodzenia."""
    osoby = list(Osoba.objects.values('imie', 'nazwisko', 'data_urodzenia'))
    KONCOWKI_K = ('a', 'A')
    imiona_m, imiona_k = Counter(), Counter()
    nazwiska = Counter()
    dekady_imiona = defaultdict(Counter)
    for o in osoby:
        im = (o['imie'] or '').strip()
        naz = (o['nazwisko'] or '').strip()
        if im:
            (imiona_k if im.endswith(KONCOWKI_K) else imiona_m)[im] += 1
            if o.get('data_urodzenia'):
                dek = (o['data_urodzenia'].year // 10) * 10
                dekady_imiona[dek][im] += 1
        if naz:
            nazwiska[naz] += 1
    top_m = imiona_m.most_common(20)
    top_k = imiona_k.most_common(20)
    top_naz = nazwiska.most_common(30)
    dekady_top = sorted(
        [(d, c.most_common(5)) for d, c in dekady_imiona.items() if c],
        key=lambda x: x[0],
    )
    return render(request, 'groby/statystyki_imion.html', {
        'top_m': top_m,
        'top_k': top_k,
        'top_nazwiska': top_naz,
        'dekady_top': dekady_top,
        'razem': len(osoby),
    })


def wydarzenia_parafialne(request):
    from django.utils import timezone
    teraz = timezone.now()
    nadchodzace = WydarzenieParafialne.objects.filter(opublikowane=True, data_start__gte=teraz).order_by('data_start')[:50]
    archiwum = WydarzenieParafialne.objects.filter(opublikowane=True, data_start__lt=teraz).order_by('-data_start')[:20]
    return render(request, 'groby/wydarzenia.html', {
        'nadchodzace': nadchodzace,
        'archiwum': archiwum,
    })


def wydarzenia_ical(request):
    """iCal feed wydarzeń parafialnych (subscribe-able)."""
    from django.utils import timezone
    teraz = timezone.now()
    qs = WydarzenieParafialne.objects.filter(opublikowane=True, data_start__gte=teraz - __import__('datetime').timedelta(days=30))
    lines = [
        'BEGIN:VCALENDAR',
        'VERSION:2.0',
        'PRODID:-//Cmentarz Szydlow//Wydarzenia//PL',
        'CALSCALE:GREGORIAN',
        'X-WR-CALNAME:Wydarzenia parafialne — Szydlow',
    ]
    for w in qs:
        koniec = w.data_koniec or w.data_start + __import__('datetime').timedelta(hours=1)
        opis = (w.opis or w.intencja or '').replace('\n', '\\n')[:500]
        lines += [
            'BEGIN:VEVENT',
            f'UID:wydarzenie-{w.pk}@szydlow',
            f'DTSTAMP:{teraz.strftime("%Y%m%dT%H%M%SZ")}',
            f'DTSTART:{w.data_start.strftime("%Y%m%dT%H%M%SZ")}',
            f'DTEND:{koniec.strftime("%Y%m%dT%H%M%SZ")}',
            f'SUMMARY:{w.tytul}',
            f'LOCATION:{w.miejsce or "Cmentarz parafialny w Szydlowie"}',
            f'DESCRIPTION:{opis}',
            'END:VEVENT',
        ]
    lines.append('END:VCALENDAR')
    return HttpResponse('\r\n'.join(lines), content_type='text/calendar; charset=utf-8')


@login_required
def eksport_notatek_pdf(request):
    """PDF z wszystkimi prywatnymi notatkami zalogowanego użytkownika (do druku/notebook)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    from django.utils import timezone
    notatki = (PrywatnaNotatka.objects.filter(user=request.user)
               .select_related('osoba__grob__sektor')
               .order_by('osoba__nazwisko', 'osoba__imie', 'data_dodania'))
    if not notatki.exists():
        messages.info(request, 'Nie masz jeszcze prywatnych notatek.')
        return redirect('groby:profil')
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title='Moje notatki')
    style = getSampleStyleSheet()
    flow = [
        Paragraph(f'Prywatne notatki — {request.user.username}', style['Title']),
        Paragraph(f'Wygenerowano: {timezone.now():%Y-%m-%d %H:%M}', style['Italic']),
        Spacer(1, 12),
    ]
    aktualna_osoba = None
    for n in notatki:
        if n.osoba_id != aktualna_osoba:
            flow.append(Spacer(1, 8))
            grob_str = f' (grób {n.osoba.grob.sektor.nazwa}/{n.osoba.grob.numer})' if n.osoba.grob_id else ''
            flow.append(Paragraph(f'<b>{n.osoba}{grob_str}</b>', style['Heading3']))
            aktualna_osoba = n.osoba_id
        flow.append(Paragraph(f'<i>{n.data_dodania:%Y-%m-%d %H:%M}</i>', style['Italic']))
        flow.append(Paragraph(n.tresc.replace('\n', '<br/>'), style['Normal']))
        flow.append(Spacer(1, 6))
    doc.build(flow)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="notatki_{request.user.username}.pdf"'
    return resp


def sezonowosc_zgonow(request):
    """Liczba zgonów per miesiąc kalendarzowy (cały okres) — sezonowość."""
    qs = Osoba.objects.exclude(data_smierci=None).values_list('data_smierci', flat=True)
    miesiace = [0] * 12
    for d in qs:
        miesiace[d.month - 1] += 1
    nazwy = ['styczeń', 'luty', 'marzec', 'kwiecień', 'maj', 'czerwiec',
             'lipiec', 'sierpień', 'wrzesień', 'październik', 'listopad', 'grudzień']
    razem = sum(miesiace) or 1
    dane = [
        {'nazwa': n, 'liczba': miesiace[i], 'procent': round(miesiace[i] * 100 / razem, 1)}
        for i, n in enumerate(nazwy)
    ]
    najwyzszy = max(miesiace) or 1
    return render(request, 'groby/sezonowosc_zgonow.html', {
        'dane': dane,
        'razem': razem,
        'najwyzszy': najwyzszy,
    })


def swiece_live(request):
    """Mapa świec aktywnych (zapalonych w ostatnich 24h)."""
    from django.utils import timezone
    from datetime import timedelta
    from django.conf import settings as dj_settings
    od = timezone.now() - timedelta(hours=24)
    swiece = (Swieca.objects.filter(data_zapalenia__gte=od, osoba__grob__plan_x__isnull=False)
              .select_related('osoba__grob__sektor')[:500])
    punkty = [
        {
            'x': s.osoba.grob.plan_x,
            'y': s.osoba.grob.plan_y,
            'osoba': str(s.osoba),
            'grob': str(s.osoba.grob),
            'osoba_id': s.osoba_id,
            'intencja': (s.intencja or '')[:200],
            'godz': s.data_zapalenia.strftime('%H:%M'),
        }
        for s in swiece if s.osoba.grob and s.osoba.grob.plan_x is not None
    ]
    if request.GET.get('format') == 'json':
        return JsonResponse({'swiece': punkty, 'liczba': len(punkty)})

    plan_url = dj_settings.MEDIA_URL + dj_settings.PLAN_IMAGE if getattr(dj_settings, 'PLAN_IMAGE', '') else None
    plan_w = plan_h = 0
    if plan_url:
        try:
            from PIL import Image
            sciezka = dj_settings.MEDIA_ROOT / dj_settings.PLAN_IMAGE
            with Image.open(sciezka) as im:
                plan_w, plan_h = im.size
        except (FileNotFoundError, OSError, ImportError):
            plan_url = None

    return render(request, 'groby/swiece_live.html', {
        'swiece_json': json.dumps(punkty),
        'liczba': len(punkty),
        'plan_url': plan_url,
        'plan_w': plan_w,
        'plan_h': plan_h,
    })


# ===== Batch 93 =====


def _wyniki_sondy(sonda):
    odpowiedzi = sonda.odpowiedzi.annotate(n=Count('glosy')).order_by('kolejnosc', 'pk')
    razem = sum(o.n for o in odpowiedzi) or 1
    return [
        {'pk': o.pk, 'tresc': o.tresc, 'n': o.n, 'procent': round(o.n * 100 / razem, 1)}
        for o in odpowiedzi
    ]


def sondy_lista(request):
    aktywne = Sonda.objects.filter(aktywna=True).prefetch_related('odpowiedzi')
    archiwum = Sonda.objects.filter(aktywna=False).prefetch_related('odpowiedzi')[:10]
    iph = _hash_ip(request)
    glosowane = set(GlosSondy.objects.filter(ip_hash=iph).values_list('odpowiedz__sonda_id', flat=True))
    return render(request, 'groby/sondy.html', {
        'aktywne': [(s, _wyniki_sondy(s), s.id in glosowane) for s in aktywne],
        'archiwum': [(s, _wyniki_sondy(s)) for s in archiwum],
    })


@require_POST
def sonda_glosuj(request, pk):
    odp = get_object_or_404(OdpowiedzSondy, pk=pk, sonda__aktywna=True)
    iph = _hash_ip(request)
    if GlosSondy.objects.filter(odpowiedz__sonda=odp.sonda, ip_hash=iph).exists():
        messages.info(request, 'Już oddałeś głos w tej sondzie.')
    else:
        try:
            GlosSondy.objects.create(
                odpowiedz=odp,
                ip_hash=iph,
                user=request.user if request.user.is_authenticated else None,
            )
            messages.success(request, 'Dziękujemy za głos!')
        except Exception:
            messages.info(request, 'Już oddałeś głos w tej sondzie.')
    return redirect('groby:sondy')


def kondolencje_lista(request):
    """Wszystkie zaakceptowane kondolencje (księga kondolencyjna)."""
    qs = (Kondolencja.objects.filter(zaakceptowana=True)
          .select_related('osoba__grob__sektor', 'autor_user')
          .order_by('-data_dodania'))
    paginator = Paginator(qs, 50)
    page = paginator.get_page(request.GET.get('p', 1))
    return render(request, 'groby/kondolencje.html', {'page': page})


def dodaj_kondolencje(request, osoba_id):
    osoba = get_object_or_404(Osoba, pk=osoba_id)
    if request.method == 'POST':
        if _antybot(request):
            return JsonResponse({'ok': False}, status=400)
        tresc = (request.POST.get('tresc') or '').strip()
        autor_imie = (request.POST.get('autor_imie') or '').strip()[:100]
        if tresc:
            Kondolencja.objects.create(
                osoba=osoba,
                autor_user=request.user if request.user.is_authenticated else None,
                autor_imie=autor_imie,
                tresc=tresc[:5000],
                zaakceptowana=False,
            )
            messages.success(request, 'Kondolencja oczekuje na akceptację moderatora. Dziękujemy.')
        return redirect('groby:osoba_detail', pk=osoba.pk)
    return render(request, 'groby/dodaj_kondolencje.html', {
        'osoba': osoba,
        'antybot_html': _pole_antybot_html(),
    })


def epitafia_galeria(request):
    """Kolekcja epitafiów — krótkich napisów z nagrobków."""
    osoby = (Osoba.objects.exclude(epitafium='')
             .select_related('grob__sektor')
             .order_by('nazwisko', 'imie'))
    return render(request, 'groby/epitafia.html', {
        'osoby': osoby,
        'liczba': osoby.count(),
    })


def zbiorki_lista(request):
    aktywne = (ZbiorkaRenowacja.objects.filter(status='aktywna')
               .select_related('grob__sektor', 'inicjator')
               .order_by('-data_zmiany'))
    zakonczone = (ZbiorkaRenowacja.objects.filter(status='zakonczona')
                  .select_related('grob__sektor')[:10])
    return render(request, 'groby/zbiorki.html', {
        'aktywne': aktywne,
        'zakonczone': zakonczone,
    })


def zbiorka_detail(request, pk):
    z = get_object_or_404(ZbiorkaRenowacja, pk=pk)
    return render(request, 'groby/zbiorka_detail.html', {'zbiorka': z})


@login_required
def zbiorka_zaproponuj(request, grob_id):
    grob = get_object_or_404(Grob, pk=grob_id)
    if request.method == 'POST':
        if _antybot(request):
            return JsonResponse({'ok': False}, status=400)
        tytul = (request.POST.get('tytul') or '').strip()[:200]
        opis = (request.POST.get('opis') or '').strip()[:5000]
        try:
            cel = max(50, min(int(request.POST.get('cel_pln') or '0'), 1_000_000))
        except ValueError:
            cel = 0
        konto = (request.POST.get('konto_bankowe') or '').strip()[:80]
        if tytul and opis and cel:
            ZbiorkaRenowacja.objects.create(
                grob=grob,
                inicjator=request.user,
                tytul=tytul,
                opis=opis,
                cel_pln=cel,
                konto_bankowe=konto,
                status='oczekuje',
            )
            messages.success(request, 'Propozycja zbiórki oczekuje na akceptację staffu.')
            return redirect('groby:grob_detail', pk=grob.pk)
        messages.error(request, 'Uzupełnij wszystkie wymagane pola.')
    return render(request, 'groby/zbiorka_form.html', {
        'grob': grob,
        'antybot_html': _pole_antybot_html(),
    })


@login_required
def najblizsza_rocznica(request):
    """JSON: najbliższa rocznica śmierci/urodzin obserwowanej osoby z odliczaniem."""
    from django.utils import timezone
    from datetime import date as dt_date
    user = request.user
    obs_qs = Osoba.objects.filter(obserwujacy__user=user).exclude(data_smierci=None)
    today = timezone.localdate()
    najblizsza = None
    najblizsze_dni = 999
    typ = ''
    for o in obs_qs:
        for d, t in [(o.data_smierci, 'rocznica śmierci'), (o.data_urodzenia, 'urodziny')]:
            if not d:
                continue
            try:
                kandydat = dt_date(today.year, d.month, d.day)
            except ValueError:
                continue
            if kandydat < today:
                try:
                    kandydat = dt_date(today.year + 1, d.month, d.day)
                except ValueError:
                    continue
            ile = (kandydat - today).days
            if ile < najblizsze_dni:
                najblizsze_dni = ile
                najblizsza = o
                typ = t
    if not najblizsza:
        return JsonResponse({'jest': False})
    return JsonResponse({
        'jest': True,
        'osoba': str(najblizsza),
        'osoba_id': najblizsza.pk,
        'typ': typ,
        'data': (today + __import__('datetime').timedelta(days=najblizsze_dni)).isoformat(),
        'dni': najblizsze_dni,
    })


def notki_cmentarne_json(request):
    """Najnowsze 5 notek mini-blogu na home (JSON dla widgetu)."""
    qs = NotkaCmentarna.objects.filter(opublikowana=True).select_related('autor')[:5]
    return JsonResponse({
        'notki': [
            {
                'id': n.pk,
                'tresc': n.tresc,
                'autor': (n.autor.username if n.autor else 'staff'),
                'data': n.data_dodania.strftime('%d.%m.%Y'),
                'przypiety': n.przypiety,
            }
            for n in qs
        ],
    })


def notki_cmentarne_lista(request):
    qs = NotkaCmentarna.objects.filter(opublikowana=True).select_related('autor')
    paginator = Paginator(qs, 30)
    page = paginator.get_page(request.GET.get('p', 1))
    return render(request, 'groby/notki.html', {'page': page})
